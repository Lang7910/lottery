# lottery.py

import requests
import json
import math
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from datetime import datetime
import os
import logging
import matplotlib.pyplot as plt
import seaborn as sns
from concurrent.futures import ThreadPoolExecutor, as_completed
from statsmodels.tsa.arima.model import ARIMA
from sklearn.ensemble import RandomForestRegressor
from sklearn.svm import SVR
import numpy as np
from statsmodels.tsa.holtwinters import ExponentialSmoothing
# 设置中文字体
import matplotlib
matplotlib.rcParams['font.sans-serif'] = ['SimHei']  # 默认字体为黑体
matplotlib.rcParams['axes.unicode_minus'] = False    # 正常显示负号
# 如果不需要 TensorFlow/keras，可注释以下行
import tensorflow as tf
from tensorflow.keras.models import Sequential
from tensorflow.keras.layers import LSTM, Dense, GRU

from sklearn.preprocessing import StandardScaler
from sklearn.cluster import KMeans

# XGBoost
from xgboost import XGBRegressor

# 线性回归 (用于 Fourier 特征示例)
from sklearn.linear_model import LinearRegression

import warnings
warnings.filterwarnings("ignore", category=FutureWarning)  # 忽略未来版本警告(可选)


# ===================== 全局常量 ====================== #
URL = "https://webapi.sporttery.cn/gateway/lottery/getHistoryPageListV1.qry?"
HEADER = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) \
AppleWebKit/537.36 (KHTML, like Gecko) Chrome/127.0.0.0 Safari/537.36",
}


# ===================== 日志配置函数 ====================== #
def setup_logging(log_file_path=None):
    """
    配置日志输出：同时输出到终端和指定日志文件。
    """
    logger = logging.getLogger()
    logger.setLevel(logging.INFO)
    
    # 防止重复添加handler
    if logger.hasHandlers():
        logger.handlers.clear()

    formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')

    # 控制台handler
    console_handler = logging.StreamHandler()
    console_handler.setLevel(logging.INFO)
    console_handler.setFormatter(formatter)
    logger.addHandler(console_handler)

    # 文件handler
    if log_file_path:
        file_handler = logging.FileHandler(log_file_path, encoding='utf-8')
        file_handler.setLevel(logging.INFO)
        file_handler.setFormatter(formatter)
        logger.addHandler(file_handler)

    return logger


# ===================== 工具函数 ====================== #
def get_input(prompt, valid_values=None):
    """封装了一个输入校验函数，用于命令行执行时使用（IDE/Notebook 中不一定有效）。"""
    while True:
        value = input(prompt)
        if valid_values and value not in valid_values:
            print("无效的输入，请重新输入")
        else:
            return value


def fetch_data(params):
    """获取指定页的数据，并返回JSON。"""
    try:
        response = requests.get(URL, headers=HEADER, params=params, timeout=10)
        response.raise_for_status()
        jsondata = response.json()
        if "value" not in jsondata or "list" not in jsondata["value"]:
            logging.error("API响应格式不正确")
            return None
        return jsondata
    except requests.exceptions.RequestException as e:
        logging.error(f"请求失败: {e}")
        return None
    except json.JSONDecodeError:
        logging.error("响应内容不是有效的JSON")
        return None


def process_page_data(page_data, data, global_index):
    """处理单页JSON数据，提取期号、号码、时间等信息并存入 data 列表。"""
    list_data = page_data["value"]["list"]
    for Num in list_data:
        DrawNum = Num["lotteryDrawNum"]
        DrawResult = Num["lotteryDrawResult"].split()
        SaleEndtime = Num["lotterySaleEndtime"]
        SaleBeginTime = Num["lotterySaleBeginTime"]
        
        front_area = list(map(int, DrawResult[:5]))
        back_area = list(map(int, DrawResult[5:]))
        
        row = [DrawNum] + front_area + back_area + [SaleBeginTime, SaleEndtime]
        data.append(row)
        logging.info(
            f"{global_index} 期号：{DrawNum} 开奖结果：{' '.join(DrawResult)} "
            f"开售时间：{SaleBeginTime} 停售时间：{SaleEndtime}"
        )
        global_index += 1
    return global_index


def create_output_folder(latest_draw_num):
    """
    创建以最新期号为名的输出文件夹，若已存在则加序号递增。
    并在主文件夹下创建若干子目录，用于分类存放输出文件。
    """
    base_dir = "outputs_lottery"
    if not os.path.exists(base_dir):
        os.makedirs(base_dir)
    
    folder_name = str(latest_draw_num)
    folder_path = os.path.join(base_dir, folder_name)
    counter = 1
    while os.path.exists(folder_path):
        folder_name = f"{latest_draw_num}_{counter}"
        folder_path = os.path.join(base_dir, folder_name)
        counter += 1
    
    # 创建主文件夹
    os.makedirs(folder_path)

    # 在主文件夹下再创建几个子文件夹
    subfolders = ["logs", "excel", "txt", "distribution_plots", "heatmaps", "trends"]
    for sub in subfolders:
        os.makedirs(os.path.join(folder_path, sub))

    return folder_path


# ==================================================================== #
#                           预测方法相关                                #
# ==================================================================== #
def moving_average_forecast(series, window=3):
    """简单移动平均。"""
    return series.rolling(window=window).mean().iloc[-1]


def arima_forecast(series, order=(5,1,0)):
    """ARIMA。"""
    try:
        model = ARIMA(series, order=order)
        model_fit = model.fit()
        forecast = model_fit.forecast()[0]
        return forecast
    except:
        return np.nan


def exponential_smoothing_forecast(series, trend=None, seasonal=None, seasonal_periods=None):
    """指数平滑（霍尔特-温特斯）。"""
    try:
        if len(series) < 2:
            return np.nan
        model = ExponentialSmoothing(series, trend=trend, seasonal=seasonal, seasonal_periods=seasonal_periods)
        model_fit = model.fit()
        forecast = model_fit.forecast(1)[0]
        return forecast
    except:
        return np.nan


def random_forest_forecast(series, n_estimators=100):
    """随机森林回归。"""
    try:
        X = np.array(range(len(series))).reshape(-1, 1)
        y = series.values
        model = RandomForestRegressor(n_estimators=n_estimators, random_state=42)
        model.fit(X, y)
        next_X = np.array([[len(series)]])
        forecast = model.predict(next_X)[0]
        return forecast
    except:
        return np.nan


def svr_forecast(series, kernel='rbf'):
    """SVR 回归。"""
    try:
        X = np.array(range(len(series))).reshape(-1, 1)
        y = series.values
        model = SVR(kernel=kernel)
        model.fit(X, y)
        next_X = np.array([[len(series)]])
        forecast = model.predict(next_X)[0]
        return forecast
    except:
        return np.nan


def lstm_forecast(series, epochs=10, batch_size=1, look_back=3):
    """简化版 LSTM 预测: 使用前 look_back 个值预测下一步。"""
    try:
        data = series.values
        if len(data) <= look_back:
            return np.nan
        
        X, y = [], []
        for i in range(len(data) - look_back):
            X.append(data[i:i+look_back])
            y.append(data[i+look_back])
        X = np.array(X)
        y = np.array(y)
        X = X.reshape((X.shape[0], X.shape[1], 1))
        
        model = Sequential()
        model.add(LSTM(20, input_shape=(look_back, 1)))
        model.add(Dense(1))
        model.compile(optimizer='adam', loss='mse')
        
        model.fit(X, y, epochs=epochs, batch_size=batch_size, verbose=0)
        
        last_sequence = data[-look_back:].reshape((1, look_back, 1))
        forecast = model.predict(last_sequence, verbose=0)[0][0]
        return forecast
    except:
        return np.nan


def mixed_model_forecast(series):
    """
    简单混合模型：ARIMA 与 LSTM 结果取平均。若有 NaN 则取另一个。
    """
    try:
        arima_pred = arima_forecast(series)
        lstm_pred = lstm_forecast(series)
        if np.isnan(arima_pred) and not np.isnan(lstm_pred):
            return lstm_pred
        elif not np.isnan(arima_pred) and np.isnan(lstm_pred):
            return arima_pred
        elif not np.isnan(arima_pred) and not np.isnan(lstm_pred):
            return (arima_pred + lstm_pred) / 2
        else:
            return np.nan
    except:
        return np.nan


# ====================== 新增的 3 种方法 ====================== #
def gru_forecast(series, epochs=10, batch_size=1, look_back=3):
    """
    使用 GRU 进行简化预测。与 LSTM 类似，只是神经元换成 GRU。
    """
    try:
        data = series.values
        if len(data) <= look_back:
            return np.nan
        
        X, y = [], []
        for i in range(len(data) - look_back):
            X.append(data[i:i+look_back])
            y.append(data[i+look_back])
        X = np.array(X)
        y = np.array(y)
        X = X.reshape((X.shape[0], X.shape[1], 1))
        
        model = Sequential()
        model.add(GRU(20, input_shape=(look_back, 1)))
        model.add(Dense(1))
        model.compile(optimizer='adam', loss='mse')
        
        model.fit(X, y, epochs=epochs, batch_size=batch_size, verbose=0)
        
        last_sequence = data[-look_back:].reshape((1, look_back, 1))
        forecast = model.predict(last_sequence, verbose=0)[0][0]
        return forecast
    except:
        return np.nan


def xgboost_forecast(series):
    """
    使用 XGBoost 进行预测，仅将索引作为特征。
    """
    try:
        X = np.array(range(len(series))).reshape(-1, 1)
        y = series.values
        model = XGBRegressor(n_estimators=100, random_state=42)
        model.fit(X, y)
        next_X = np.array([[len(series)]])
        forecast = model.predict(next_X)[0]
        return forecast
    except:
        return np.nan


def fourier_lr_forecast(series, k=3):
    """
    简易 Fourier 特征 + 线性回归。仅作示例。
    """
    try:
        data = series.values
        n = len(data)
        if n < 2:
            return np.nan
        
        # 做 FFT 并只取前 k 个最显著的正频率
        fft_vals = np.fft.fft(data)
        freqs = np.fft.fftfreq(n)

        positive_indices = np.where(freqs > 0)
        freqs_pos = freqs[positive_indices]
        fft_pos_vals = fft_vals[positive_indices]

        amps = np.abs(fft_pos_vals)            
        top_k_idx = np.argsort(amps)[::-1][:k] 

        selected_freqs = freqs_pos[top_k_idx]
        selected_amps = amps[top_k_idx]
        selected_phases = np.angle(fft_pos_vals[top_k_idx])

        # 构造特征
        feature_matrix = []
        for i in range(n):
            row = [i]
            for f, amp, ph in zip(selected_freqs, selected_amps, selected_phases):
                row.append(amp * np.sin(2 * np.pi * f * i + ph))
                row.append(amp * np.cos(2 * np.pi * f * i + ph))
            feature_matrix.append(row)
        feature_matrix = np.array(feature_matrix)

        lr = LinearRegression()
        lr.fit(feature_matrix, data)

        # 预测下一个点
        next_row = [n]
        for f, amp, ph in zip(selected_freqs, selected_amps, selected_phases):
            next_row.append(amp * np.sin(2 * np.pi * f * n + ph))
            next_row.append(amp * np.cos(2 * np.pi * f * n + ph))
        next_row = np.array(next_row).reshape(1, -1)

        forecast = lr.predict(next_row)[0]
        return forecast
    except:
        return np.nan


def get_forecasts(series):
    """统一获取各种模型的预测结果。"""
    forecasts = {}
    # 原先已有的方法
    forecasts['Moving_Average'] = moving_average_forecast(series)
    forecasts['ARIMA'] = arima_forecast(series)
    forecasts['Random_Forest'] = random_forest_forecast(series)
    forecasts['SVR'] = svr_forecast(series)
    forecasts['LSTM'] = lstm_forecast(series)
    forecasts['Exponential_Smoothing'] = exponential_smoothing_forecast(series)
    forecasts['Mixed_Model'] = mixed_model_forecast(series)
    
    # 新增方法
    forecasts['GRU'] = gru_forecast(series)
    forecasts['XGBoost'] = xgboost_forecast(series)
    forecasts['Fourier_LR'] = fourier_lr_forecast(series)

    return forecasts


# ==================================================================== #
#                        统计 & 可视化相关函数                          #
# ==================================================================== #
def compute_statistics(data):
    """
    统计前区1-5、后区1-2 的出现次数和概率，并用一列 '区' 标识(front/back)。
    """
    total_draws = len(data)
    # 前区统计容器
    front_counts = {num: {pos: 0 for pos in range(1,6)} for num in range(1,36)}
    # 后区统计容器
    back_counts = {num: {pos: 0 for pos in range(1,3)} for num in range(1,13)}
    
    # 累计次数
    for row in data:
        front = row[1:6]  # 前区5个
        back = row[6:8]   # 后区2个
        for pos, num in enumerate(front, start=1):
            front_counts[num][pos] += 1
        for pos, num in enumerate(back, start=1):
            back_counts[num][pos] += 1
    
    stats_rows = []
    
    # 前区 1~35
    for num in range(1,36):
        row = [num, "front"]
        for pos in range(1,6):
            count = front_counts[num][pos]
            prob = count / total_draws
            row.append(count)
            row.append(prob)
        # 后区占位：4个
        row.extend([0, 0, 0, 0])
        stats_rows.append(row)
    
    # 后区 1~12
    for num in range(1,13):
        row = [num, "back"]
        row.extend([0, 0]*5)
        for pos in range(1,3):
            count = back_counts[num][pos]
            prob = count / total_draws
            row.append(count)
            row.append(prob)
        stats_rows.append(row)
    
    columns = ["号码","区"]
    for pos in range(1,6):
        columns.append(f"前区{pos}_出现次数")
        columns.append(f"前区{pos}_概率")
    for pos in range(1,3):
        columns.append(f"后区{pos}_出现次数")
        columns.append(f"后区{pos}_概率")
    
    stats_df = pd.DataFrame(stats_rows, columns=columns)
    return stats_df


def apply_color_to_stats(ws, stats_df, start_row, start_col):
    """
    给统计结果上色：出现次数最高的金色、未出现的灰色，其余正常。
    """
    max_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")  # 金色
    never_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")  # 灰色
    normal_fill = PatternFill(fill_type=None)
    
    # 前区1~5
    for pos in range(1,6):
        col_name = f"前区{pos}_出现次数"
        max_count = stats_df[col_name].max()
        count_col = start_col + (pos-1)*2 + 1
        for idx, row in stats_df.iterrows():
            cell = ws.cell(row=start_row + idx, column=count_col)
            val = row[col_name]
            if val == max_count and max_count != 0:
                cell.fill = max_fill
            elif val == 0:
                cell.fill = never_fill
            else:
                cell.fill = normal_fill
    
    # 后区1~2
    for pos in range(1,3):
        col_name = f"后区{pos}_出现次数"
        count_col = start_col + 10 + (pos-1)*2 + 1
        max_count = stats_df[col_name].max()
        for idx, row in stats_df.iterrows():
            cell = ws.cell(row=start_row + idx, column=count_col)
            val = row[col_name]
            if val == max_count and max_count != 0:
                cell.fill = max_fill
            elif val == 0:
                cell.fill = never_fill
            else:
                cell.fill = normal_fill


def plot_distribution(df, title, xlabel, ylabel, output_path):
    """绘制分布柱状图，并保存到指定路径。"""
    plt.figure(figsize=(16,8))
    df['号码'] = pd.to_numeric(df['号码'], errors='coerce')
    df['出现次数'] = pd.to_numeric(df['出现次数'], errors='coerce')
    sns.barplot(x='号码', y='出现次数', hue='位置', data=df)
    plt.title(title, fontsize=16)
    plt.xlabel(xlabel, fontsize=14)
    plt.ylabel(ylabel, fontsize=14)
    plt.legend(title='位置')
    plt.xticks(fontsize=12)
    plt.yticks(fontsize=12)
    plt.tight_layout()
    plt.savefig(output_path)
    plt.close()
    logging.info(f"已保存图表: {output_path}")


def plot_heatmap(stats_df, positions, num_range, title, output_path, front=True):
    """
    绘制热力图 (前区 or 后区)，并保存到指定路径。
    """
    try:
        if front:
            # 只取区=front，并且号码在 1~35
            sub_df = stats_df[(stats_df["区"] == "front") & (stats_df["号码"].between(1, 35))]
            value_vars = [f"前区{pos}_出现次数" for pos in range(1,6)]
            label_prefix = "前区"
        else:
            # 只取区=back，并且号码在 1~12
            sub_df = stats_df[(stats_df["区"] == "back") & (stats_df["号码"].between(1, 12))]
            value_vars = [f"后区{pos}_出现次数" for pos in range(1,3)]
            label_prefix = "后区"
        
        melted = sub_df.melt(
            id_vars=["号码"],
            value_vars=value_vars,
            var_name="位置",
            value_name="出现次数"
        )
        
        melted["位置"] = melted["位置"].str.extract(rf'{label_prefix}(\d+)_').astype(int)
        pivot_table = melted.pivot(index="位置", columns="号码", values="出现次数")
        
        plt.figure(figsize=(num_range/2.5, positions + 2))
        sns.heatmap(
            pivot_table, 
            annot=True, 
            fmt=".0f", 
            cmap="YlGnBu", 
            annot_kws={"size": 8}
        )
        plt.title(title, fontsize=16)
        plt.xlabel('号码', fontsize=14)
        plt.ylabel('位置', fontsize=14)
        plt.xticks(fontsize=10)
        plt.yticks(fontsize=12, rotation=0)
        plt.tight_layout()
        plt.savefig(output_path)
        plt.close()
        logging.info(f"已保存热力图: {output_path}")
    except Exception as e:
        logging.error(f"绘制热力图时发生错误: {e}")


def plot_trend(data, title, xlabel, ylabel, output_path):
    """
    通用的趋势图绘制，data 应该包含列：['期号','位置','号码']。
    """
    try:
        data['号码'] = pd.to_numeric(data['号码'], errors='coerce')
        data['期号'] = data['期号'].astype(str)
        
        plt.figure(figsize=(18,9))
        sns.lineplot(data=data, x='期号', y='号码', hue='位置', marker='o')
        plt.title(title, fontsize=16)
        plt.xlabel(xlabel, fontsize=14)
        plt.ylabel(ylabel, fontsize=14)
        plt.legend(title='位置')
        plt.xticks(rotation=45, fontsize=10)
        plt.yticks(fontsize=12)
        plt.tight_layout()
        plt.savefig(output_path)
        plt.close()
        logging.info(f"已保存走势图: {output_path}")
    except Exception as e:
        logging.error(f"绘制走势图时发生错误: {e}")


def plot_trend_with_prediction(df, prediction_dict, title_prefix, output_path_prefix):
    """
    针对某个模型生成的预测号码，与最近15期一起画前区/后区走势图。
    df: 原 DataFrame，包含期号、前区/后区列
    prediction_dict: {pos_name: pred_value, ...}
    output_path_prefix: 用于拼接输出文件路径 (不含 .png 扩展名)
    """
    latest_15 = df.tail(15).copy()
    predict_row = {'期号': 'predict'}
    for pos in ['前区1','前区2','前区3','前区4','前区5','后区1','后区2']:
        predict_row[pos] = prediction_dict.get(pos, np.nan)
    
    new_row_df = pd.DataFrame([predict_row])
    latest_15 = pd.concat([latest_15, new_row_df], ignore_index=True)

    # 前区
    front_melt = latest_15.melt(
        id_vars=["期号"],
        value_vars=["前区1","前区2","前区3","前区4","前区5"],
        var_name="位置",
        value_name="号码"
    )
    front_out = f"{output_path_prefix}_front.png"
    plot_trend(front_melt, f"{title_prefix}_前区趋势", "期号", "号码", front_out)

    # 后区
    back_melt = latest_15.melt(
        id_vars=["期号"],
        value_vars=["后区1","后区2"],
        var_name="位置",
        value_name="号码"
    )
    back_out = f"{output_path_prefix}_back.png"
    plot_trend(back_melt, f"{title_prefix}_后区趋势", "期号", "号码", back_out)


# ==================================================================== #
#                               主流程                                 #
# ==================================================================== #
def main():
    # ============ 第 1 步：用户选择查询方式 =============
    query_type = get_input("1: 按期数查询\n2: 按期号查询\n请选择查询方式: ", ['1', '2'])
    
    params = {
        "gameNo": "85",
        "provinceId": "0",
        "startTerm": "",
        "endTerm": "",
        "pageNo": "1",
        "pageSize": "30",
        "isVerify": "1",
        "termLimits": ""
    }
    
    if query_type == '1':
        term_limits = get_input("最近100期即输入100,最多近一年\n请输入期数：")
        while not term_limits.isdigit() or not (1 <= int(term_limits) <= 100):
            print("请输入一个介于1到100之间的有效期数。")
            term_limits = get_input("请输入期数：")
        params["termLimits"] = term_limits
    elif query_type == '2':
        print("期号格式如yyxxx, 如25001，最多支持5年")
        start_term = input("请输入起始期号yyxxx：")
        end_term = input("请输入结束期号yyxxx：")
        while not (start_term.isdigit() and end_term.isdigit()):
            print("期号应为数字，请重新输入。")
            start_term = input("请输入起始期号yyxxx：")
            end_term = input("请输入结束期号yyxxx：")
        params["startTerm"] = start_term
        params["endTerm"] = end_term
    
    # ============ 第 2 步：获取第一页数据 ============
    global_index = 1
    data = []
    jsondata = fetch_data(params)
    if not jsondata:
        logging.error("无法获取数据，程序终止")
        return
    
    pages = jsondata["value"]["pages"]
    pageNo = jsondata["value"]["pageNo"]
    pageSize = jsondata["value"]["pageSize"]
    total = jsondata["value"]["total"]
    
    global_index = process_page_data(jsondata, data, global_index)
    
    # 如果总页数大于1，则多线程获取剩余页面
    if pages > 1:
        with ThreadPoolExecutor(max_workers=5) as executor:
            futures = {
                executor.submit(fetch_data, {**params, "pageNo": str(page)}): page
                for page in range(2, pages + 1)
            }
            results = []
            for future in as_completed(futures):
                page = futures[future]
                jsondata_page = future.result()
                if jsondata_page:
                    results.append((page, jsondata_page))
                else:
                    logging.warning(f"第{page}页数据获取失败，跳过")
        
        # 按页号排序再处理
        results_sorted = sorted(results, key=lambda x: x[0])
        for page, jsondata_page in results_sorted:
            global_index = process_page_data(jsondata_page, data, global_index)
    
    logging.info(f"页数：{pages}")
    logging.info(f"第{pageNo}页")
    logging.info(f"页大小：{pageSize}")
    logging.info(f"总期数：{total}")
    
    # ============ 第 3 步：将数据写入 DataFrame =============
    columns = ["期号","前区1","前区2","前区3","前区4","前区5","后区1","后区2","开售时间","停售时间"]
    df = pd.DataFrame(data, columns=columns)
    latest_draw_num = df.iloc[0]["期号"]  # 通常第一行是最新一期
    
    # ============ 第 4 步：创建输出文件夹并初始化日志 ============
    folder_path = create_output_folder(latest_draw_num)
    current_time = datetime.now().strftime("%Y%m%d_%H%M%S")

    # 日志文件路径
    log_file_path = os.path.join(folder_path, "logs", f"run_{current_time}.log")
    setup_logging(log_file_path=log_file_path)
    logging.info("开始执行主流程...")

    # ============ 第 5 步：保存为 Excel =============
    wb = Workbook()
    ws = wb.active
    ws.title = "Lottery Results"
    ws.append(columns)

    front_fill_odd = PatternFill(start_color="4682B4", end_color="4682B4", fill_type="solid")
    front_fill_even = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
    back_fill_odd = PatternFill(start_color="32CD32", end_color="32CD32", fill_type="solid")
    back_fill_even = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
    
    for row_idx, row_val in enumerate(data, start=2):
        ws.append(row_val)
        # 前区上色
        for i in range(1, 6):
            cell = ws.cell(row=row_idx, column=i+1)
            if row_idx % 2 == 0:
                cell.fill = front_fill_even
            else:
                cell.fill = front_fill_odd
        # 后区上色
        for i in range(6, 8):
            cell = ws.cell(row=row_idx, column=i+1)
            if row_idx % 2 == 0:
                cell.fill = back_fill_even
            else:
                cell.fill = back_fill_odd

    stats_df = compute_statistics(data)
    ws_stats = wb.create_sheet(title="统计结果")
    ws_stats.append(stats_df.columns.tolist())
    for index, row_val in stats_df.iterrows():
        ws_stats.append(row_val.tolist())
    apply_color_to_stats(ws_stats, stats_df, start_row=2, start_col=3)

    excel_dir = os.path.join(folder_path, "excel")
    excel_filepath = os.path.join(excel_dir, f"lottery_results_{current_time}.xlsx")
    wb.save(excel_filepath)
    logging.info(f"数据已保存到 {excel_filepath}")

    # ============ 第 6 步：数据可视化 =============
    distribution_dir = os.path.join(folder_path, "distribution_plots")
    heatmap_dir = os.path.join(folder_path, "heatmaps")
    trend_dir = os.path.join(folder_path, "trends")

    # (1) 前区 / 后区分布柱状图
    front_distribution_df = stats_df[
        (stats_df["区"] == "front") & (stats_df["号码"].between(1,35))
    ].melt(
        id_vars=["号码","区"],
        value_vars=[f"前区{pos}_出现次数" for pos in range(1,6)],
        var_name="位置",
        value_name="出现次数"
    )
    front_distribution_df["位置"] = front_distribution_df["位置"].str.extract(r'前区(\d)_').astype(int)

    back_distribution_df = stats_df[
        (stats_df["区"] == "back") & (stats_df["号码"].between(1,12))
    ].melt(
        id_vars=["号码","区"],
        value_vars=[f"后区{pos}_出现次数" for pos in range(1,3)],
        var_name="位置",
        value_name="出现次数"
    )
    back_distribution_df["位置"] = back_distribution_df["位置"].str.extract(r'后区(\d)_').astype(int)

    front_dist_png = os.path.join(distribution_dir, f"front_distribution_{current_time}.png")
    back_dist_png = os.path.join(distribution_dir, f"back_distribution_{current_time}.png")

    plot_distribution(front_distribution_df, "前区号码分布", "号码", "出现次数", front_dist_png)
    plot_distribution(back_distribution_df,  "后区号码分布", "号码", "出现次数", back_dist_png)

    # (2) 热力图（前区 / 后区）
    front_heatmap_png = os.path.join(heatmap_dir, f"front_heatmap_{current_time}.png")
    plot_heatmap(stats_df, positions=5, num_range=35,
                 title="前区号码热力图",
                 output_path=front_heatmap_png,
                 front=True)

    back_heatmap_png = os.path.join(heatmap_dir, f"back_heatmap_{current_time}.png")
    plot_heatmap(stats_df, positions=2, num_range=12,
                 title="后区号码热力图",
                 output_path=back_heatmap_png,
                 front=False)

    # (3) 走势图 (最近15期)
    latest_15 = df.tail(15).copy()
    latest_15_front = latest_15.melt(
        id_vars=["期号"],
        value_vars=["前区1","前区2","前区3","前区4","前区5"],
        var_name="位置",
        value_name="号码"
    )
    front_trend_png = os.path.join(trend_dir, f"trend_front_{current_time}.png")
    plot_trend(latest_15_front, "最近15期前区号码走势图", "期号", "号码", front_trend_png)

    latest_15_back = latest_15.melt(
        id_vars=["期号"],
        value_vars=["后区1","后区2"],
        var_name="位置",
        value_name="号码"
    )
    back_trend_png = os.path.join(trend_dir, f"trend_back_{current_time}.png")
    plot_trend(latest_15_back, "最近15期后区号码走势图", "期号", "号码", back_trend_png)

    # ============ 第 7 步：预测与聚类 =============
    positions_list = ['前区1','前区2','前区3','前区4','前区5','后区1','后区2']
    all_forecasts = {}
    
    methods = [
        'Moving_Average','ARIMA','Random_Forest','SVR',
        'LSTM','Exponential_Smoothing','Mixed_Model',
        'GRU','XGBoost','Fourier_LR'
    ]

    logging.info("开始对每个位置进行多模型预测...")
    for pos in positions_list:
        series = df[pos]
        all_forecasts[pos] = get_forecasts(series)
        for method in methods:
            logging.info(f"[预测细节] 位置 {pos} - 模型 {method} -> {all_forecasts[pos][method]}")

    # 将预测结果合并到 DataFrame
    forecast_df = pd.DataFrame(index=methods, columns=positions_list)
    for pos in positions_list:
        for method in methods:
            forecast_df.at[method, pos] = all_forecasts[pos][method]

    forecast_df = forecast_df.astype(float)
    # 备份原始值
    forecast_df_raw = forecast_df.copy()
    forecast_df = forecast_df.fillna(forecast_df.mean())

    # 同步填充后的值回 all_forecasts
    for method in forecast_df.index:
        for pos in positions_list:
            all_forecasts[pos][method] = forecast_df.at[method, pos]

    # 聚类
    scaler = StandardScaler()
    scaled_data = scaler.fit_transform(forecast_df)
    kmeans = KMeans(n_clusters=5, random_state=42)
    clusters = kmeans.fit_predict(scaled_data)
    forecast_df['Cluster'] = clusters
    logging.info("预测方法与聚类结果：\n" + str(forecast_df))

    # 根据聚类结果，挑选每组的第一个方法为代表
    final_number_sets = []
    for cluster in range(5):
        cluster_methods = forecast_df[forecast_df['Cluster'] == cluster].index
        if len(cluster_methods) == 0:
            continue
        representative_method = cluster_methods[0]
        number_set = {}
        for pos in positions_list:
            number_set[pos] = all_forecasts[pos][representative_method]
        final_number_sets.append((representative_method, number_set))

    # ============ 第 8 步：输出结果到TXT =============
    try:
        predict_term = str(int(latest_draw_num) + 1)
    except:
        predict_term = "NextTerm"

    txt_dir = os.path.join(folder_path, "txt")
    txt_filename = os.path.join(txt_dir, f"prediction_results_{current_time}.txt")
    used_data_count = len(df)

    with open(txt_filename, "w", encoding="utf-8") as f:
        f.write(f"预测期号：{predict_term}\n")
        f.write(f"使用数据期数：{used_data_count}\n\n")

        print("----------------------")
        print(f"预测期号：{predict_term}")
        print(f"使用数据期数：{used_data_count}")
        print("----------------------")

        for method in methods:
            predicted_positions = []
            for pos in positions_list:
                val = all_forecasts[pos][method]
                if math.isnan(val):
                    logging.warning(f"方法 {method} 在位置 {pos} 的预测结果是 NaN，使用 0 替代")
                    val = 0
                predicted_positions.append(int(round(val)))

            f.write(f"方法 {method} 的预测号码：{predicted_positions}\n")
            logging.info(f"方法 {method} 的预测号码：{predicted_positions}")
            print(f"方法 {method} 的预测号码：{predicted_positions}")

            # 画上带预测的走势图
            pred_dict = dict(zip(positions_list, predicted_positions))
            pred_plot_prefix = os.path.join(trend_dir, f"{method}_predict_{current_time}")
            plot_trend_with_prediction(df, pred_dict, title_prefix=f"{method}_预测", output_path_prefix=pred_plot_prefix)

        f.write("\n====== 聚类后的号码组合 ======\n")
        print("\n====== 聚类后的号码组合 ======")

        for i, (rep_method, number_set) in enumerate(final_number_sets, start=1):
            numbers = []
            for pos in positions_list:
                val = number_set[pos]
                if math.isnan(val):
                    logging.warning(f"聚类组 {i}, 方法 {rep_method} 在位置 {pos} 出现 NaN，使用 0 替代")
                    val = 0
                numbers.append(int(round(val)))
            f.write(f"聚类组 {i}（代表方法：{rep_method}） -> {numbers}\n")
            logging.info(f"聚类组 {i}（代表方法：{rep_method}） -> {numbers}")
            print(f"聚类组 {i}（代表方法：{rep_method}） -> {numbers}")

    logging.info(f"\n预测结果已写入：{txt_filename}")
    print(f"\n预测结果已写入：{txt_filename}")


if __name__ == "__main__":
    main()
