# time_series_prediction_with_all.py
import matplotlib
matplotlib.use('Agg')  # 设置非交互式后端

import os
import sys
import time
import random
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt

# scikit-learn
from sklearn.ensemble import RandomForestRegressor
from sklearn.svm import SVR
from sklearn.linear_model import BayesianRidge

# statsmodels
from statsmodels.tsa.holtwinters import SimpleExpSmoothing
from statsmodels.tsa.arima.model import ARIMA

# tensorflow (LSTM)
import tensorflow as tf
from tensorflow.keras.models import Sequential
from tensorflow.keras.layers import LSTM, Dense

# kmodes for clustering
try:
    from kmodes.kmodes import KModes
except ImportError:
    print("请先安装 kmodes 模块: pip install kmodes")
    sys.exit(1)

from collections import Counter

# OpenPyXL for Excel operations
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.styles import PatternFill

import logging
from pathlib import Path

import argparse  # 导入 argparse
from concurrent.futures import ThreadPoolExecutor, as_completed  # 引入并发模块

# 设置命令行参数解析
def parse_arguments():
    parser = argparse.ArgumentParser(description="时间序列预测脚本")
    parser.add_argument("excel_file", help="Excel文件路径")
    parser.add_argument("output_dir", nargs="?", default=None, help="输出目录路径")
    parser.add_argument("--no-kill", action="store_true", help="不应用蓝球杀号方法")
    return parser.parse_args()

# 解析命令行参数
args = parse_arguments()

# 设置日志
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.StreamHandler(sys.stdout),  # 输出到终端
        logging.FileHandler("prediction.log", encoding='utf-8')  # 同时输出到日志文件
    ]
)

# 设置随机种子
random.seed(42)
np.random.seed(42)
tf.random.set_seed(42)

# 定义项目根目录
PROJECT_ROOT = Path(__file__).parent.resolve()

# 定义输出目录
if args.output_dir:
    OUTPUT_DIR = Path(args.output_dir)
else:
    OUTPUT_DIR = PROJECT_ROOT / "prediction_results"

# 创建输出目录（如果未创建）
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

# 定义输出文件路径
details_file_path = OUTPUT_DIR / "prediction_details.txt"
summary_file_path = OUTPUT_DIR / "prediction_summary.txt"

# 定义预测详情目录和文件
PREDICTIONS_DIR = OUTPUT_DIR / "predictions"
DETAILS_DIR = PREDICTIONS_DIR / "details"
RESULTS_DIR = PREDICTIONS_DIR / "results"

# 创建目录
PREDICTIONS_DIR.mkdir(parents=True, exist_ok=True)
DETAILS_DIR.mkdir(parents=True, exist_ok=True)
RESULTS_DIR.mkdir(parents=True, exist_ok=True)

# 定义预测详情文件路径
preview_details_path = DETAILS_DIR / f"preview详情.txt"

# 创建输出文件（如果不存在）
if not details_file_path.exists():
    with open(details_file_path, 'w', encoding='utf-8') as f:
        f.write("预测详情\n")

if not summary_file_path.exists():
    with open(summary_file_path, 'w', encoding='utf-8') as f:
        f.write("预测摘要\n")

# 设置中文字体（若需要）
plt.rcParams['font.sans-serif'] = ['SimHei']
plt.rcParams['axes.unicode_minus'] = False

# 定义输出目录
CHARTS_DIR = OUTPUT_DIR / "charts"
RED_PLOTS_DIR = CHARTS_DIR / "red_plots"
BLUE_PLOTS_DIR = CHARTS_DIR / "blue_plots"
PREDICTIONS_DIR = OUTPUT_DIR / "predictions"
DETAILS_DIR = PREDICTIONS_DIR / "details"
RESULTS_DIR = PREDICTIONS_DIR / "results"

# 创建所有必要的目录（如果未创建）
for directory in [RED_PLOTS_DIR, BLUE_PLOTS_DIR, DETAILS_DIR, RESULTS_DIR]:
    directory.mkdir(parents=True, exist_ok=True)

# 定义文本文件的路径
details_file_path = DETAILS_DIR / "preview详情.txt"
results_file_path = RESULTS_DIR / "preview结果.txt"

# 定义预测方法参数
DEFAULT_N_LAGS = 5
DEFAULT_EPOCHS = 10

def moving_average_prediction(series, window=5):
    try:
        if len(series) < window:
            return series.iloc[-1]
        return series.iloc[-window:].mean()
    except Exception as e:
        logging.error(f"移动平均预测失败: {e}", exc_info=True)
        return float(series.iloc[-1])

def exponential_smoothing_prediction(series, alpha=0.3):
    try:
        model = SimpleExpSmoothing(series.astype(float)).fit(smoothing_level=alpha, optimized=False)
        forecast = model.forecast(1)
        return forecast.iloc[0]
    except Exception as e:
        logging.error(f"指数平滑预测失败: {e}", exc_info=True)
        return float(series.iloc[-1])

def create_supervised_data(values, n_lags=5):
    X, y = [], []
    for i in range(len(values) - n_lags):
        X.append(values[i:i+n_lags])
        y.append(values[i+n_lags])
    return np.array(X), np.array(y)

def ml_prediction(series, method="rf", n_lags=5):
    try:
        s = series.values if hasattr(series, 'values') else np.array(series)
        X, Y = create_supervised_data(s, n_lags)
        if len(X) < 1:
            return float(s[-1])

        if method == "rf":
            model = RandomForestRegressor(n_estimators=100, random_state=42)
        elif method == "svr":
            model = SVR()
        elif method == "bayes":
            model = BayesianRidge()
        else:
            raise ValueError("Unknown ML method in ml_prediction()")

        model.fit(X, Y)
        last_data = s[-n_lags:]
        predicted = model.predict([last_data])
        return float(predicted[0])
    except Exception as e:
        logging.error(f"机器学习预测失败 ({method}): {e}", exc_info=True)
        return float(series.iloc[-1])

def arima_prediction(series, p=1, d=0, q=1):
    try:
        s = series.values if hasattr(series, 'values') else np.array(series)
        if len(s) < max(p, d, q) + 1:
            return float(s[-1])

        model = ARIMA(s, order=(p, d, q))
        model_fit = model.fit()
        forecast = model_fit.forecast(steps=1)
        return float(forecast[0])
    except Exception as e:
        logging.error(f"ARIMA预测失败: {e}", exc_info=True)
        return float(series.iloc[-1])

def lstm_prediction(series, n_lags=5, epochs=10):
    try:
        s = series.values if hasattr(series, 'values') else np.array(series)
        X, Y = create_supervised_data(s, n_lags)
        if len(X) < 1:
            return float(s[-1])

        X = X.reshape(X.shape[0], X.shape[1], 1)
        Y = Y.reshape(-1, 1)

        model = Sequential()
        model.add(LSTM(units=20, input_shape=(n_lags, 1)))
        model.add(Dense(1))

        model.compile(optimizer='adam', loss='mse')
        model.fit(X, Y, epochs=epochs, batch_size=1, verbose=0)

        last_data = s[-n_lags:].reshape((1, n_lags, 1))
        pred = model.predict(last_data)
        return float(pred[0][0])
    except Exception as e:
        logging.error(f"LSTM预测失败: {e}", exc_info=True)
        return float(series.iloc[-1])

def hybrid_prediction(series, weights={'arima': 0.6, 'lstm': 0.4}):
    try:
        pred_arima = arima_prediction(series, p=1, d=0, q=1)
        pred_lstm  = lstm_prediction(series, n_lags=5, epochs=10)
        return weights['arima'] * pred_arima + weights['lstm'] * pred_lstm
    except Exception as e:
        logging.error(f"混合模型预测失败: {e}", exc_info=True)
        return float(series.iloc[-1])

def predict_next_number(series, method="ma"):
    try:
        if method == "ma":
            return moving_average_prediction(series, window=5)
        elif method == "es":
            return exponential_smoothing_prediction(series, alpha=0.3)
        elif method in ["rf", "svr", "bayes"]:
            return ml_prediction(series, method=method, n_lags=5)
        elif method == "arima":
            return arima_prediction(series, p=1, d=0, q=1)
        elif method == "lstm":
            return lstm_prediction(series, n_lags=5, epochs=10)
        elif method == "hybrid":
            return hybrid_prediction(series)
        else:
            raise ValueError(f"Unsupported method: {method}")
    except Exception as e:
        logging.error(f"预测方法 {method} 失败: {e}", exc_info=True)
        return float(series.iloc[-1])

def plot_blue_last10(periods, last10, new_value, new_period, method="ma"):
    try:
        display_periods = [str(p)[-3:] for p in periods]
        display_new_period = str(new_period)[-3:]

        # 输出用于绘图的期号和蓝球号码
        logging.debug(f"绘图蓝球期号: {display_periods}")
        logging.debug(f"绘图新蓝球期号: {display_new_period}")
        logging.debug(f"绘图蓝球号码数据: {last10}")

        # 创建 Figure 和 Axes 对象
        fig, ax = plt.subplots(figsize=(12, 6))

        # 绘制历史数据
        ax.plot(display_periods, last10, marker='o', color='b', label='历史数据')

        # 绘制预测值
        ax.plot(display_new_period, new_value, marker='D', color='r', label='新预测')

        # 设置标题和标签
        ax.set_title(f"蓝球走势（{method.upper()}）")
        ax.set_xlabel("期号")
        ax.set_ylabel("蓝球号码")
        plt.xticks(rotation=45)
        ax.legend()
        plt.tight_layout()

        # 保存图像
        time_str = time.strftime("%Y%m%d_%H%M%S")
        filename = BLUE_PLOTS_DIR / f"蓝球走势_{method.upper()}_{time_str}.png"
        fig.savefig(filename, dpi=100)
        plt.close(fig)

        logging.info(f"蓝球走势图已保存到 {filename}")
    except Exception as e:
        logging.error(f"绘制蓝球走势图失败 ({method}): {e}", exc_info=True)

def plot_red_positions_in_one_figure(df, predictions, method="ma", lookback=30, periods=None):
    try:
        display_periods = [str(p)[-3:] for p in periods]
        display_new_period = str(max(periods) + 1)[-3:]

        # 输出用于绘图的期号和红球号码
        logging.debug(f"绘图红球期号: {display_periods}")
        logging.debug(f"绘图新红球期号: {display_new_period}")
        for pos in range(1, 7):
            col_name = f"红球号码{pos}"
            series_pos = df[col_name].tail(lookback).iloc[-10:]
            logging.debug(f"绘图红球位置 {pos} 数据: {series_pos.tolist()}")

        # 创建 Figure 和 Axes 对象
        fig, ax = plt.subplots(figsize=(14, 8))

        colors = ["blue", "green", "orange", "purple", "brown", "grey"]
        labels = [f"红{pos}历史" for pos in range(1, 7)]

        for pos in range(1, 7):
            col_name = f"红球号码{pos}"
            series_pos = df[col_name].tail(lookback).iloc[-10:]
            last10_pos = series_pos.tolist()

            # 绘制历史数据
            ax.plot(display_periods, last10_pos, marker='o', color=colors[pos-1], label=labels[pos-1] if pos ==1 else "")

            # 绘制预测值
            ax.plot(display_new_period, predictions[pos], marker='D', color=colors[pos-1])

        # 设置标题和标签
        ax.set_title(f"红球 6 个位置走势（{method.upper()}）")
        ax.set_xlabel("期号")
        ax.set_ylabel("红球号码")
        plt.xticks(rotation=45)

        # 创建图例
        handles, labels = ax.get_legend_handles_labels()
        by_label = dict(zip(labels, handles))
        ax.legend(by_label.values(), by_label.keys(), loc='best')

        plt.tight_layout()

        # 保存图像
        time_str = time.strftime("%Y%m%d_%H%M%S")
        filename = RED_PLOTS_DIR / f"红球走势_{method.upper()}_{time_str}.png"
        fig.savefig(filename, dpi=100)
        plt.close(fig)

        logging.info(f"红球走势图已保存到 {filename}")
    except Exception as e:
        logging.error(f"绘制红球走势图失败 ({method}): {e}", exc_info=True)

def recommend_by_method(df, method="ma", lookback=None, apply_blue_kill=True):
    if len(df) < 15:
        return [f"数据不足15期，{method.upper()} 方法暂不执行。"], None

    # 使用传递的 lookback 参数，如果未传递，则使用全部数据
    lookback = len(df) if lookback is None else lookback
    logging.info(f"方法 {method.upper()} 使用的 lookback 值: {lookback}")

    # 提取用于预测的蓝球数据
    blue_series = df["蓝球号码"].tail(lookback)
    logging.debug(f"蓝球数据 (最近 {lookback} 期):\n{blue_series}")

    blue_pred_float = predict_next_number(blue_series, method=method)
    blue_int = max(1, min(16, int(round(blue_pred_float))))

    # 如果启用了蓝球杀号，应用杀号规则
    if apply_blue_kill and len(df) >= 2:
        # 获取最新两期的蓝球号码
        latest_blue = df["蓝球号码"].iloc[-1]
        prev_blue = df["蓝球号码"].iloc[-2] if len(df) >= 2 else None
        
        # 计算各种杀号方法的结果
        killed_numbers = set()
        
        # 方法1: 15-上期蓝球尾数
        kill_digit1 = (15 - (latest_blue % 10)) % 10
        kill_numbers1 = [kill_digit1, kill_digit1 + 10] if kill_digit1 != 0 else [10]
        kill_numbers1 = [int(n) for n in kill_numbers1 if 1 <= n <= 16]
        killed_numbers.update(kill_numbers1)
        
        # 方法2: 19-上期蓝球尾数
        kill_digit2 = (19 - (latest_blue % 10)) % 10
        kill_numbers2 = [kill_digit2, kill_digit2 + 10] if kill_digit2 != 0 else [10]
        kill_numbers2 = [int(n) for n in kill_numbers2 if 1 <= n <= 16]
        killed_numbers.update(kill_numbers2)
        
        # 方法3: 21-上期蓝球尾数
        kill_digit3 = (21 - (latest_blue % 10)) % 10
        kill_numbers3 = [kill_digit3, kill_digit3 + 10] if kill_digit3 != 0 else [10]
        kill_numbers3 = [int(n) for n in kill_numbers3 if 1 <= n <= 16]
        killed_numbers.update(kill_numbers3)
        
        if prev_blue is not None:
            # 方法4: 上两期蓝号的头和尾相加
            head_prev = prev_blue // 10
            tail_latest = latest_blue % 10
            kill_digit4 = (head_prev + tail_latest) % 10
            kill_numbers4 = [kill_digit4, kill_digit4 + 10] if kill_digit4 != 0 else [10]
            kill_numbers4 = [int(n) for n in kill_numbers4 if 1 <= n <= 16]
            killed_numbers.update(kill_numbers4)
            
            # 方法5: 上两期蓝号的尾和头相加
            tail_prev = prev_blue % 10
            head_latest = latest_blue // 10
            kill_digit5 = (tail_prev + head_latest) % 10
            kill_numbers5 = [kill_digit5, kill_digit5 + 10] if kill_digit5 != 0 else [10]
            kill_numbers5 = [int(n) for n in kill_numbers5 if 1 <= n <= 16]
            killed_numbers.update(kill_numbers5)
            
            # 方法6: 上两期蓝号的尾相加
            tail_prev = prev_blue % 10
            tail_latest = latest_blue % 10
            kill_digit6 = (tail_prev + tail_latest) % 10
            kill_numbers6 = [kill_digit6, kill_digit6 + 10] if kill_digit6 != 0 else [10]
            kill_numbers6 = [int(n) for n in kill_numbers6 if 1 <= n <= 16]
            killed_numbers.update(kill_numbers6)
        
        # 如果预测的蓝球在被杀号列表中，选择一个未被杀的号码
        if blue_int in killed_numbers:
            available_numbers = [n for n in range(1, 17) if n not in killed_numbers]
            if available_numbers:
                # 找到与原预测值最接近的未被杀的号码
                blue_int = min(available_numbers, key=lambda x: abs(x - blue_int))
            # 如果所有号码都被杀了（极少发生），保持原预测值

    last10_blue = list(blue_series.iloc[-10:])
    last10_periods = list(df["期数"].tail(10))

    new_period = df.iloc[-1]["期数"] + 1

    print_buffer = [
        f"\n========== 使用 {method.upper()} 预测 ==========",
        f"【{method.upper()}】蓝球最近10期：{last10_blue}",
        f"【{method.upper()}】蓝球新预测（浮点）: {blue_pred_float:.2f}",
        f"【{method.upper()}】蓝球取整后: {blue_int}"
    ]
    
    if apply_blue_kill:
        print_buffer.append(f"【{method.upper()}】蓝球杀号列表: {sorted(list(killed_numbers))}")
        print_buffer.append(f"【{method.upper()}】蓝球最终预测(考虑杀号): {blue_int}")

    plot_blue_last10(last10_periods, last10_blue, blue_int, new_period, method=method)

    # 计算红球杀号方法的结果（成功率高于70%的方法）
    red_killed_numbers = set()
    if len(df) >= 2:
        # 获取最新一期的数据
        latest_idx = len(df) - 1
        
        # 获取最新一期的红球
        latest_red_balls = []
        for j in range(1, 7):
            col_name = f"红球号码{j}"
            latest_red_balls.append(df.iloc[latest_idx][col_name])
        
        latest_blue = df["蓝球号码"].iloc[latest_idx]
        
        # 判断一个数是否为质数的函数
        def is_prime(n):
            if n <= 1:
                return False
            if n <= 3:
                return True
            if n % 2 == 0 or n % 3 == 0:
                return False
            i = 5
            while i * i <= n:
                if n % i == 0 or n % (i + 2) == 0:
                    return False
                i += 6
            return True
        
        # 计算红球中质数的个数
        def count_primes(red_balls):
            return sum(1 for ball in red_balls if is_prime(ball))
        
        # 计算红球尾数之和
        def sum_of_tails(red_balls):
            return sum(ball % 10 for ball in red_balls)
        
        # 计算红球之和的各位数字之和
        def sum_of_digits(n):
            return sum(int(digit) for digit in str(n))
        
        # 对称码计算（以33为中心的对称）
        def symmetric_code(n):
            return 34 - n  # 33+1-n
        
        # 计算AC值（红球间隔种类数）
        def calculate_ac_value(red_balls):
            # 对红球排序
            sorted_balls = sorted(red_balls)
            # 计算相邻红球的间隔
            gaps = [sorted_balls[i+1] - sorted_balls[i] for i in range(len(sorted_balls)-1)]
            # 计算间隔的种类数（AC值）
            unique_gaps = len(set(gaps))
            return unique_gaps
        
        # 计算红球杀号方法的成功率
        _, red_success_rates, _ = red_kill_methods(df)
        
        # 只应用成功率高于70%的方法
        high_success_methods = [int(method.replace("方法", "")) for method, rate in red_success_rates.items() if rate > 0.7]
        
        if 2 in high_success_methods:  # 方法2: 红号尾数之和杀号法
            tail_sum = sum_of_tails(latest_red_balls)
            for multiple in range(1, 4):
                num = tail_sum * multiple
                if 1 <= num <= 33:
                    red_killed_numbers.add(num)
        
        if 3 in high_success_methods:  # 方法3: 质数个数杀号法
            prime_count = count_primes(latest_red_balls)
            for multiple in range(1, 6):
                num = prime_count * multiple
                if 1 <= num <= 33:
                    red_killed_numbers.add(num)
        
        if 4 in high_success_methods:  # 方法4: 红号第一位对称码杀号法
            first_red = latest_red_balls[0]
            sym_code = symmetric_code(first_red)
            red_killed_numbers.add(sym_code)
            if 1 <= sym_code + 1 <= 33:
                red_killed_numbers.add(sym_code + 1)
            if 1 <= sym_code - 1 <= 33:
                red_killed_numbers.add(sym_code - 1)
        
        if 5 in high_success_methods:  # 方法5: 红号第一位+第六位杀号法
            first_plus_last = latest_red_balls[0] + latest_red_balls[5]
            if 1 <= first_plus_last <= 33:
                red_killed_numbers.add(first_plus_last)
            half_sum = first_plus_last // 2
            if 1 <= half_sum <= 33:
                red_killed_numbers.add(half_sum)
        
        if 6 in high_success_methods:  # 方法6: 出球顺序第一位+第二位+第五位杀号法
            sum_125 = latest_red_balls[0] + latest_red_balls[1] + latest_red_balls[4]
            if sum_125 > 33:
                num = sum_125 // 2
                if 1 <= num <= 33:
                    red_killed_numbers.add(num)
            else:
                red_killed_numbers.add(sum_125)
        
        if 1 in high_success_methods:  # 方法1: 蓝号杀号法
            for multiple in range(1, 4):
                num = latest_blue * multiple
                if 1 <= num <= 33:
                    red_killed_numbers.add(num)
        
        if 7 in high_success_methods:  # 方法7: 红号之和的和杀号法
            red_sum = sum(latest_red_balls)
            digit_sum = sum_of_digits(red_sum)
            for multiple in range(1, 4):
                num = digit_sum * multiple
                if 1 <= num <= 33:
                    red_killed_numbers.add(num)
        
        if 8 in high_success_methods:  # 方法8: AC值杀号法
            ac_value = calculate_ac_value(latest_red_balls)
            if 1 <= ac_value <= 2:
                for num in [1, 4, 7, 10, 13, 16, 19, 22, 25, 28, 31]:
                    if 1 <= num <= 33:
                        red_killed_numbers.add(num)
            elif 3 <= ac_value <= 4:
                for num in [2, 5, 8, 11, 14, 17, 20, 23, 26, 29, 32]:
                    if 1 <= num <= 33:
                        red_killed_numbers.add(num)
            else:  # AC值 >= 5
                for num in [3, 6, 9, 12, 15, 18, 21, 24, 27, 30, 33]:
                    if 1 <= num <= 33:
                        red_killed_numbers.add(num)
        
        if 9 in high_success_methods:  # 方法9: 蓝号+AC值杀号法
            blue_ac_sum = latest_blue + ac_value
            if blue_ac_sum <= 10:
                for num in [1, 5, 9, 13, 17, 21, 25, 29, 33]:
                    if 1 <= num <= 33:
                        red_killed_numbers.add(num)
            elif 11 <= blue_ac_sum <= 20:
                for num in [2, 6, 10, 14, 18, 22, 26, 30]:
                    if 1 <= num <= 33:
                        red_killed_numbers.add(num)
            else:  # blue_ac_sum > 20
                for num in [3, 7, 11, 15, 19, 23, 27, 31]:
                    if 1 <= num <= 33:
                        red_killed_numbers.add(num)
        
        if 10 in high_success_methods:  # 方法10: 蓝号+质数个数杀号法
            blue_prime_sum = latest_blue + prime_count
            if blue_prime_sum % 3 == 0:
                for num in [3, 6, 9, 12, 15, 18, 21, 24, 27, 30, 33]:
                    if 1 <= num <= 33:
                        red_killed_numbers.add(num)
            elif blue_prime_sum % 3 == 1:
                for num in [1, 4, 7, 10, 13, 16, 19, 22, 25, 28, 31]:
                    if 1 <= num <= 33:
                        red_killed_numbers.add(num)
            else:  # blue_prime_sum % 3 == 2
                for num in [2, 5, 8, 11, 14, 17, 20, 23, 26, 29, 32]:
                    if 1 <= num <= 33:
                        red_killed_numbers.add(num)
        
        if 11 in high_success_methods:  # 方法11: AC值×质数个数杀号法
            ac_prime_product = ac_value * prime_count
            if 1 <= ac_prime_product <= 33:
                red_killed_numbers.add(ac_prime_product)
            for multiple in range(2, 4):
                num = ac_prime_product * multiple
                if 1 <= num <= 33:
                    red_killed_numbers.add(num)
        
        if 12 in high_success_methods:  # 方法12: AC值+质数个数杀号法
            ac_prime_sum = ac_value + prime_count
            if 1 <= ac_prime_sum <= 33:
                red_killed_numbers.add(ac_prime_sum)
            for multiple in range(2, 4):
                num = ac_prime_sum * multiple
                if 1 <= num <= 33:
                    red_killed_numbers.add(num)
        
        if 13 in high_success_methods:  # 方法13: AC值-红号第六位杀号法
            sixth_red = latest_red_balls[5]
            ac_sixth_diff = abs(ac_value - sixth_red)
            if 1 <= ac_sixth_diff <= 33:
                red_killed_numbers.add(ac_sixth_diff)
        
        if 14 in high_success_methods:  # 方法14: 蓝号-红号第一、六位杀号法
            first_red = latest_red_balls[0]
            blue_first_diff = abs(latest_blue - first_red)
            blue_sixth_diff = abs(latest_blue - sixth_red)
            if 1 <= blue_first_diff <= 33:
                red_killed_numbers.add(blue_first_diff)
            if 1 <= blue_sixth_diff <= 33 and blue_sixth_diff != blue_first_diff:
                red_killed_numbers.add(blue_sixth_diff)
        
        if 15 in high_success_methods:  # 方法15: 蓝号+AC值+质数个数杀号法
            blue_ac_prime_sum = latest_blue + ac_value + prime_count
            if 1 <= blue_ac_prime_sum <= 33:
                red_killed_numbers.add(blue_ac_prime_sum)
            else:
                remainder = blue_ac_prime_sum % 33
                if remainder > 0:
                    red_killed_numbers.add(remainder)
                else:
                    red_killed_numbers.add(33)
        
        if 16 in high_success_methods:  # 方法16: 红号×0.88杀号法
            for red in latest_red_balls:
                num = int(red * 0.88)
                if 1 <= num <= 33:
                    red_killed_numbers.add(num)
        
        if 17 in high_success_methods:  # 方法17: 红号第三位对称码+7杀号法
            third_red = latest_red_balls[2]
            third_sym_plus7 = symmetric_code(third_red) + 7
            if 1 <= third_sym_plus7 <= 33:
                red_killed_numbers.add(third_sym_plus7)

    # 预测红球
    red_predictions = {}
    red_vals = []
    for pos in range(1, 7):
        col_name = f"红球号码{pos}"
        series_pos = df[col_name].tail(lookback)
        logging.debug(f"红球位置 {pos} 数据 (最近 {lookback} 期):\n{series_pos}")

        pred_val_float = predict_next_number(series_pos, method=method)
        pred_int = max(1, min(33, int(round(pred_val_float))))
        
        # 如果预测的红球在被杀号列表中，选择一个未被杀的号码
        if pred_int in red_killed_numbers:
            available_numbers = [n for n in range(1, 34) if n not in red_killed_numbers and n not in red_vals]
            if available_numbers:
                # 找到与原预测值最接近的未被杀的号码
                pred_int = min(available_numbers, key=lambda x: abs(x - pred_int))

        last10_pos = list(series_pos.iloc[-10:])
        print_buffer.append(f"【{method.upper()}】红{pos}最近10期：{last10_pos}")
        print_buffer.append(
            f"【{method.upper()}】红{pos}新预测（浮点）: {pred_val_float:.2f}, 取整后: {pred_int}"
        )

        red_predictions[pos] = pred_int
        red_vals.append(pred_int)

    if red_killed_numbers:
        print_buffer.append(f"【{method.upper()}】红球杀号列表: {sorted(list(red_killed_numbers))}")

    plot_red_positions_in_one_figure(df, predictions=red_predictions, method=method, lookback=lookback, periods=last10_periods)

    # 去重+补足+排序
    reds_unique = sorted(set(red_vals))
    if len(reds_unique) < 6:
        # 排除被杀的号码
        remaining = list(set(range(1, 34)) - set(reds_unique) - red_killed_numbers)
        if not remaining:  # 如果所有号码都被杀了，则使用所有未使用的号码
            remaining = list(set(range(1, 34)) - set(reds_unique))
        additional = random.sample(remaining, min(6 - len(reds_unique), len(remaining)))
        reds_unique.extend(additional)
        logging.debug(f"红球补充的额外号码: {additional}")
    reds_unique = sorted(reds_unique[:6])

    rec_str = f"【{method.upper()}】推荐号码：红球{reds_unique} + 蓝球{blue_int}"
    print_buffer.append(rec_str)

    return print_buffer, (reds_unique, blue_int)

def generate_recommendations(results_map, num_recommendations=5):
    red_sets = []
    blue_counts = []
    for method, (buf, res) in results_map.items():
        if res is not None:
            reds, blue = res
            red_sets.append(reds)
            blue_counts.append(blue)

    if not red_sets:
        logging.warning("没有可用的红球预测数据。")
        # 返回一些默认的推荐，避免后续处理出错
        default_recommendations = []
        for i in range(min(num_recommendations, 5)):
            default_reds = sorted(random.sample(range(1, 34), 6))
            default_blue = random.randint(1, 16)
            default_recommendations.append((default_reds, default_blue))
        return default_recommendations

    def set_to_binary_vector(s, n=33):
        vec = [0]*n
        for num in s:
            vec[num-1] = 1
        return vec

    binary_vectors = [set_to_binary_vector(s) for s in red_sets]
    binary_vectors = np.array(binary_vectors)

    try:
        km = KModes(n_clusters=num_recommendations, init='Huang', n_init=5, verbose=0)
        clusters = km.fit_predict(binary_vectors)
    except Exception as e:
        logging.error(f"K-Modes 聚类失败: {e}", exc_info=True)
        return []

    recommendations = []
    for cluster_num in range(num_recommendations):
        cluster_indices = [i for i, c in enumerate(clusters) if c == cluster_num]
        if not cluster_indices:
            continue
        cluster_reds = [tuple(sorted(red_sets[i])) for i in cluster_indices]
        all_red_numbers = [num for sublist in cluster_reds for num in sublist]
        red_counter = Counter(all_red_numbers)
        representative_reds = sorted([num for num, cnt in red_counter.most_common(6)])

        cluster_blues = [blue_counts[i] for i in cluster_indices]
        blue_counter = Counter(cluster_blues)
        most_common_blue = blue_counter.most_common(1)[0][0]

        recommendations.append((representative_reds, most_common_blue))

    while len(recommendations) < num_recommendations:
        all_red_numbers = [num for s in red_sets for num in s]
        red_counter = Counter(all_red_numbers)
        top_reds = sorted([num for num, cnt in red_counter.most_common(6)])

        blue_counter = Counter(blue_counts)
        top_blue = blue_counter.most_common(1)[0][0]

        if (top_reds, top_blue) not in recommendations:
            recommendations.append((top_reds, top_blue))
        else:
            break

    return recommendations

# 蓝球杀号方法
def blue_kill_methods(df):
    """
    实现多种蓝球杀号方法并计算其成功率
    
    方法1: 用15减去上期蓝球号码，得出的数就是下期要杀的蓝号尾数
    方法2: 用19减上期蓝号得出的数即为下期要杀的尾数
    方法3: 用21减上期蓝号得出的数就是下期要杀的尾数
    方法4: 用上两期蓝号的头和尾相加的数即为下期要杀的蓝号尾数
    方法5: 用上两期蓝号的尾和头相加的数即为下期要杀的尾数
    方法6: 用上二期蓝号的尾相加得出的数就是下期要杀的尾数
    
    返回:
        - 各方法的杀号结果
        - 各方法的历史成功率
        - 最新一期的杀号预测
    """
    # 确保数据按期数升序排序
    df = df.sort_values(by="期数").reset_index(drop=True)
    
    # 提取蓝球列
    blue_balls = df["蓝球号码"].values
    periods = df["期数"].values
    
    # 初始化结果存储
    results = {
        "期数": periods[1:],  # 从第二期开始才能有杀号结果
        "蓝球号码": blue_balls[1:],
        "方法1_杀号": [],
        "方法1_结果": [],
        "方法2_杀号": [],
        "方法2_结果": [],
        "方法3_杀号": [],
        "方法3_结果": [],
        "方法4_杀号": [],
        "方法4_结果": [],
        "方法5_杀号": [],
        "方法5_结果": [],
        "方法6_杀号": [],
        "方法6_结果": [],
    }
    
    # 方法1-3只需要上一期数据
    for i in range(1, len(blue_balls)):
        prev_blue = blue_balls[i-1]
        current_blue = blue_balls[i]
        
        # 方法1: 15-上期蓝球尾数
        kill_digit1 = 15 - (prev_blue % 10)
        kill_digit1 = kill_digit1 % 10  # 确保是个位数
        kill_numbers1 = [kill_digit1, kill_digit1 + 10] if kill_digit1 != 0 else [10]
        kill_numbers1 = [int(n) for n in kill_numbers1 if 1 <= n <= 16]  # 确保在1-16范围内
        results["方法1_杀号"].append(kill_numbers1)
        results["方法1_结果"].append(current_blue not in kill_numbers1)
        
        # 方法2: 19-上期蓝球尾数
        kill_digit2 = 19 - (prev_blue % 10)
        kill_digit2 = kill_digit2 % 10  # 确保是个位数
        kill_numbers2 = [kill_digit2, kill_digit2 + 10] if kill_digit2 != 0 else [10]
        kill_numbers2 = [int(n) for n in kill_numbers2 if 1 <= n <= 16]  # 确保在1-16范围内
        results["方法2_杀号"].append(kill_numbers2)
        results["方法2_结果"].append(current_blue not in kill_numbers2)
        
        # 方法3: 21-上期蓝球尾数
        kill_digit3 = 21 - (prev_blue % 10)
        kill_digit3 = kill_digit3 % 10  # 确保是个位数
        kill_numbers3 = [kill_digit3, kill_digit3 + 10] if kill_digit3 != 0 else [10]
        kill_numbers3 = [int(n) for n in kill_numbers3 if 1 <= n <= 16]  # 确保在1-16范围内
        results["方法3_杀号"].append(kill_numbers3)
        results["方法3_结果"].append(current_blue not in kill_numbers3)
    
    # 方法4-6需要上两期数据
    for i in range(2, len(blue_balls)):
        prev_blue1 = blue_balls[i-2]  # 上上期
        prev_blue2 = blue_balls[i-1]  # 上期
        current_blue = blue_balls[i]  # 当期
        
        # 方法4: 上两期蓝号的头和尾相加
        head1 = prev_blue1 // 10
        tail2 = prev_blue2 % 10
        kill_digit4 = (head1 + tail2) % 10
        kill_numbers4 = [kill_digit4, kill_digit4 + 10] if kill_digit4 != 0 else [10]
        kill_numbers4 = [int(n) for n in kill_numbers4 if 1 <= n <= 16]  # 确保在1-16范围内
        
        # 方法5: 上两期蓝号的尾和头相加
        tail1 = prev_blue1 % 10
        head2 = prev_blue2 // 10
        kill_digit5 = (tail1 + head2) % 10
        kill_numbers5 = [kill_digit5, kill_digit5 + 10] if kill_digit5 != 0 else [10]
        kill_numbers5 = [int(n) for n in kill_numbers5 if 1 <= n <= 16]  # 确保在1-16范围内
        
        # 方法6: 上两期蓝号的尾相加
        tail1 = prev_blue1 % 10
        tail2 = prev_blue2 % 10
        kill_digit6 = (tail1 + tail2) % 10
        kill_numbers6 = [kill_digit6, kill_digit6 + 10] if kill_digit6 != 0 else [10]
        kill_numbers6 = [int(n) for n in kill_numbers6 if 1 <= n <= 16]  # 确保在1-16范围内
        
        # 填充方法4-6的结果
        if i == 2:  # 为第一个有方法4-6结果的期数填充前面的空值
            results["方法4_杀号"].append([])
            results["方法4_结果"].append(None)
            results["方法5_杀号"].append([])
            results["方法5_结果"].append(None)
            results["方法6_杀号"].append([])
            results["方法6_结果"].append(None)
        
        results["方法4_杀号"].append(kill_numbers4)
        results["方法4_结果"].append(current_blue not in kill_numbers4)
        
        results["方法5_杀号"].append(kill_numbers5)
        results["方法5_结果"].append(current_blue not in kill_numbers5)
        
        results["方法6_杀号"].append(kill_numbers6)
        results["方法6_结果"].append(current_blue not in kill_numbers6)
    
    # 计算各方法的成功率
    success_rates = {}
    for method in range(1, 7):
        method_key = f"方法{method}_结果"
        # 过滤掉None值
        valid_results = [r for r in results[method_key] if r is not None]
        if valid_results:
            success_count = sum(valid_results)
            total_count = len(valid_results)
            success_rate = success_count / total_count
            success_rates[f"方法{method}"] = success_rate
    
    # 生成最新一期的杀号预测
    latest_blue = blue_balls[-1]
    if len(blue_balls) >= 2:
        prev_blue = blue_balls[-2]
        
        # 方法1-3的最新预测
        next_kill1 = [(15 - (latest_blue % 10)) % 10]
        if next_kill1[0] != 0:
            next_kill1.append(next_kill1[0] + 10)
        next_kill1 = [int(n) for n in next_kill1 if 1 <= n <= 16]
        
        next_kill2 = [(19 - (latest_blue % 10)) % 10]
        if next_kill2[0] != 0:
            next_kill2.append(next_kill2[0] + 10)
        next_kill2 = [int(n) for n in next_kill2 if 1 <= n <= 16]
        
        next_kill3 = [(21 - (latest_blue % 10)) % 10]
        if next_kill3[0] != 0:
            next_kill3.append(next_kill3[0] + 10)
        next_kill3 = [int(n) for n in next_kill3 if 1 <= n <= 16]
        
        # 方法4-6的最新预测
        head_prev = prev_blue // 10
        tail_prev = prev_blue % 10
        head_latest = latest_blue // 10
        tail_latest = latest_blue % 10
        
        next_kill4 = [(head_prev + tail_latest) % 10]
        if next_kill4[0] != 0:
            next_kill4.append(next_kill4[0] + 10)
        next_kill4 = [int(n) for n in next_kill4 if 1 <= n <= 16]
        
        next_kill5 = [(tail_prev + head_latest) % 10]
        if next_kill5[0] != 0:
            next_kill5.append(next_kill5[0] + 10)
        next_kill5 = [int(n) for n in next_kill5 if 1 <= n <= 16]
        
        next_kill6 = [(tail_prev + tail_latest) % 10]
        if next_kill6[0] != 0:
            next_kill6.append(next_kill6[0] + 10)
        next_kill6 = [int(n) for n in next_kill6 if 1 <= n <= 16]
        
        next_predictions = {
            "方法1": next_kill1,
            "方法2": next_kill2,
            "方法3": next_kill3,
            "方法4": next_kill4,
            "方法5": next_kill5,
            "方法6": next_kill6
        }
    else:
        next_predictions = {}
    
    # 创建结果DataFrame
    results_df = pd.DataFrame({
        "期数": results["期数"],
        "蓝球号码": results["蓝球号码"],
        "方法1_杀号": [','.join(map(str, x)) for x in results["方法1_杀号"]],
        "方法1_结果": results["方法1_结果"],
        "方法2_杀号": [','.join(map(str, x)) for x in results["方法2_杀号"]],
        "方法2_结果": results["方法2_结果"],
        "方法3_杀号": [','.join(map(str, x)) for x in results["方法3_杀号"]],
        "方法3_结果": results["方法3_结果"],
        "方法4_杀号": [','.join(map(str, x)) if x else '' for x in results["方法4_杀号"]],
        "方法4_结果": results["方法4_结果"],
        "方法5_杀号": [','.join(map(str, x)) if x else '' for x in results["方法5_杀号"]],
        "方法5_结果": results["方法5_结果"],
        "方法6_杀号": [','.join(map(str, x)) if x else '' for x in results["方法6_杀号"]],
        "方法6_结果": results["方法6_结果"],
    })
    
    return results_df, success_rates, next_predictions

# 将蓝球杀号结果写入Excel
def write_blue_kill_results_to_excel(excel_file, results_df, success_rates, next_predictions):
    """
    将蓝球杀号结果写入Excel文件的新工作表
    """
    try:
        # 加载工作簿
        workbook = load_workbook(excel_file)
        
        # 检查是否已存在蓝球杀号工作表，如果存在则删除
        if "蓝球杀号" in workbook.sheetnames:
            del workbook["蓝球杀号"]
        
        # 创建新的工作表
        sheet = workbook.create_sheet("蓝球杀号")
        
        # 定义填充颜色
        success_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")  # 绿色
        fail_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")     # 红色
        
        # 写入标题行
        headers = ["期数", "蓝球号码", 
                  "方法1杀号", "方法1结果", 
                  "方法2杀号", "方法2结果", 
                  "方法3杀号", "方法3结果", 
                  "方法4杀号", "方法4结果", 
                  "方法5杀号", "方法5结果", 
                  "方法6杀号", "方法6结果"]
        for col, header in enumerate(headers, 1):
            sheet.cell(row=1, column=col).value = header
        
        # 写入数据
        for row_idx, row in results_df.iterrows():
            sheet.cell(row=row_idx+2, column=1).value = row["期数"]
            sheet.cell(row=row_idx+2, column=2).value = row["蓝球号码"]
            
            sheet.cell(row=row_idx+2, column=3).value = row["方法1_杀号"]
            cell = sheet.cell(row=row_idx+2, column=4)
            cell.value = "成功" if row["方法1_结果"] else "失败" if row["方法1_结果"] is not None else ""
            if row["方法1_结果"] is not None:
                cell.fill = success_fill if row["方法1_结果"] else fail_fill
            
            sheet.cell(row=row_idx+2, column=5).value = row["方法2_杀号"]
            cell = sheet.cell(row=row_idx+2, column=6)
            cell.value = "成功" if row["方法2_结果"] else "失败" if row["方法2_结果"] is not None else ""
            if row["方法2_结果"] is not None:
                cell.fill = success_fill if row["方法2_结果"] else fail_fill
            
            sheet.cell(row=row_idx+2, column=7).value = row["方法3_杀号"]
            cell = sheet.cell(row=row_idx+2, column=8)
            cell.value = "成功" if row["方法3_结果"] else "失败" if row["方法3_结果"] is not None else ""
            if row["方法3_结果"] is not None:
                cell.fill = success_fill if row["方法3_结果"] else fail_fill
            
            sheet.cell(row=row_idx+2, column=9).value = row["方法4_杀号"]
            cell = sheet.cell(row=row_idx+2, column=10)
            cell.value = "成功" if row["方法4_结果"] else "失败" if row["方法4_结果"] is not None else ""
            if row["方法4_结果"] is not None:
                cell.fill = success_fill if row["方法4_结果"] else fail_fill
            
            sheet.cell(row=row_idx+2, column=11).value = row["方法5_杀号"]
            cell = sheet.cell(row=row_idx+2, column=12)
            cell.value = "成功" if row["方法5_结果"] else "失败" if row["方法5_结果"] is not None else ""
            if row["方法5_结果"] is not None:
                cell.fill = success_fill if row["方法5_结果"] else fail_fill
            
            sheet.cell(row=row_idx+2, column=13).value = row["方法6_杀号"]
            cell = sheet.cell(row=row_idx+2, column=14)
            cell.value = "成功" if row["方法6_结果"] else "失败" if row["方法6_结果"] is not None else ""
            if row["方法6_结果"] is not None:
                cell.fill = success_fill if row["方法6_结果"] else fail_fill
        
        # 写入成功率统计
        row_offset = len(results_df) + 3
        sheet.cell(row=row_offset, column=1).value = "杀号方法成功率统计"
        for i, (method, rate) in enumerate(success_rates.items(), 1):
            sheet.cell(row=row_offset+i, column=1).value = method
            cell = sheet.cell(row=row_offset+i, column=2)
            cell.value = f"{rate:.2%}"
            # 根据成功率设置颜色
            if rate >= 0.8:
                cell.fill = success_fill
            elif rate < 0.5:
                cell.fill = fail_fill
        
        # 写入下一期预测
        row_offset = row_offset + len(success_rates) + 2
        next_period = results_df["期数"].iloc[-1] + 1
        sheet.cell(row=row_offset, column=1).value = f"{next_period}期蓝球杀号预测"
        for i, (method, numbers) in enumerate(next_predictions.items(), 1):
            sheet.cell(row=row_offset+i, column=1).value = method
            sheet.cell(row=row_offset+i, column=2).value = ','.join(map(str, numbers))
        
        # 保存工作簿
        workbook.save(excel_file)
        logging.info(f"蓝球杀号结果已写入Excel文件 '{excel_file}' 的 '蓝球杀号' 工作表")
        return True
    except Exception as e:
        logging.error(f"写入Excel时发生错误: {e}")
        return False

def red_kill_methods(df):
    """
    实现多种红球杀号方法并计算其成功率
    
    方法1: 蓝号杀号法 - 根据蓝球号码杀掉特定的红球号码
    方法2: 红号尾数之和杀号法 - 根据上期红球尾数之和杀号
    方法3: 质数个数杀号法 - 根据上期红球中质数的个数杀号
    方法4: 红号第一位对称码杀号法 - 根据上期红球第一位的对称码杀号
    方法5: 红号第一位+第六位杀号法 - 根据上期红球第一位和第六位之和杀号
    方法6: 出球顺序第一位+第二位+第五位杀号法 - 根据上期红球出球顺序的特定位置之和杀号
    方法7: 红号之和的和杀号法 - 根据上期红球之和的各位数字之和杀号
    方法8: AC值杀号法
    方法9: 蓝号+AC值杀号法
    方法10: 蓝号+质数个数杀号法
    方法11: AC值×质数个数杀号法
    方法12: AC值+质数个数杀号法
    方法13: AC值-红号第六位杀号法
    方法14: 蓝号-红号第一、六位杀号法
    方法15: 蓝号+AC值+质数个数杀号法
    方法16: 红号×0.88杀号法
    方法17: 红号第三位对称码+7杀号法
    
    返回:
        - 各方法的杀号结果
        - 各方法的历史成功率
        - 最新一期的杀号预测
    """
    # 确保数据按期数升序排序
    df = df.sort_values(by="期数").reset_index(drop=True)
    
    # 提取红球列和蓝球列
    periods = df["期数"].values
    blue_balls = df["蓝球号码"].values
    
    # 初始化结果存储
    results = {
        "期数": periods[1:],  # 从第二期开始才能有杀号结果
        "方法1_杀号": [],
        "方法1_结果": [],
        "方法2_杀号": [],
        "方法2_结果": [],
        "方法3_杀号": [],
        "方法3_结果": [],
        "方法4_杀号": [],
        "方法4_结果": [],
        "方法5_杀号": [],
        "方法5_结果": [],
        "方法6_杀号": [],
        "方法6_结果": [],
        "方法7_杀号": [],
        "方法7_结果": [],
        "方法8_杀号": [],  # AC值杀号法
        "方法8_结果": [],
        "方法9_杀号": [],  # 蓝号+AC值杀号法
        "方法9_结果": [],
        "方法10_杀号": [],  # 蓝号+质数个数杀号法
        "方法10_结果": [],
        "方法11_杀号": [],  # AC值×质数个数杀号法
        "方法11_结果": [],
        "方法12_杀号": [],  # AC值+质数个数杀号法
        "方法12_结果": [],
        "方法13_杀号": [],  # AC值-红号第六位杀号法
        "方法13_结果": [],
        "方法14_杀号": [],  # 蓝号-红号第一、六位杀号法
        "方法14_结果": [],
        "方法15_杀号": [],  # 蓝号+AC值+质数个数杀号法
        "方法15_结果": [],
        "方法16_杀号": [],  # 红号×0.88杀号法
        "方法16_结果": [],
        "方法17_杀号": [],  # 红号第三位对称码+7杀号法
        "方法17_结果": [],
    }
    
    # 判断一个数是否为质数的函数
    def is_prime(n):
        if n <= 1:
            return False
        if n <= 3:
            return True
        if n % 2 == 0 or n % 3 == 0:
            return False
        i = 5
        while i * i <= n:
            if n % i == 0 or n % (i + 2) == 0:
                return False
            i += 6
        return True
    
    # 计算红球中质数的个数
    def count_primes(red_balls):
        return sum(1 for ball in red_balls if is_prime(ball))
    
    # 计算红球尾数之和
    def sum_of_tails(red_balls):
        return sum(ball % 10 for ball in red_balls)
    
    # 计算红球之和的各位数字之和
    def sum_of_digits(n):
        return sum(int(digit) for digit in str(n))
    
    # 对称码计算（以33为中心的对称）
    def symmetric_code(n):
        return 34 - n  # 33+1-n
    
    # 计算AC值（红球间隔种类数）
    def calculate_ac_value(red_balls):
        # 对红球排序
        sorted_balls = sorted(red_balls)
        # 计算相邻红球的间隔
        gaps = [sorted_balls[i+1] - sorted_balls[i] for i in range(len(sorted_balls)-1)]
        # 计算间隔的种类数（AC值）
        unique_gaps = len(set(gaps))
        return unique_gaps
    
    for i in range(1, len(periods)):
        # 获取上期数据
        prev_red_balls = []
        for j in range(1, 7):
            col_name = f"红球号码{j}"
            prev_red_balls.append(df.iloc[i-1][col_name])
        
        prev_blue = blue_balls[i-1]
        
        # 获取当期数据
        current_red_balls = []
        for j in range(1, 7):
            col_name = f"红球号码{j}"
            current_red_balls.append(df.iloc[i][col_name])
        
        # 方法1: 蓝号杀号法 - 根据蓝球号码杀掉特定的红球号码
        # 这里我们简单地杀掉与蓝球号码相关的红球：蓝球号码及其倍数
        kill_numbers1 = []
        for multiple in range(1, 4):  # 杀掉蓝球号码的1-3倍
            num = prev_blue * multiple
            if 1 <= num <= 33:
                kill_numbers1.append(num)
        
        results["方法1_杀号"].append(kill_numbers1)
        results["方法1_结果"].append(all(num not in current_red_balls for num in kill_numbers1))
        
        # 方法2: 红号尾数之和杀号法
        tail_sum = sum_of_tails(prev_red_balls)
        kill_numbers2 = []
        # 杀掉尾数之和的倍数
        for multiple in range(1, 4):
            num = tail_sum * multiple
            if 1 <= num <= 33:
                kill_numbers2.append(num)
        
        results["方法2_杀号"].append(kill_numbers2)
        results["方法2_结果"].append(all(num not in current_red_balls for num in kill_numbers2))
        
        # 方法3: 质数个数杀号法
        prime_count = count_primes(prev_red_balls)
        kill_numbers3 = []
        # 杀掉质数个数的倍数
        for multiple in range(1, 6):
            num = prime_count * multiple
            if 1 <= num <= 33:
                kill_numbers3.append(num)
        
        results["方法3_杀号"].append(kill_numbers3)
        results["方法3_结果"].append(all(num not in current_red_balls for num in kill_numbers3))
        
        # 方法4: 红号第一位对称码杀号法
        first_red = prev_red_balls[0]
        sym_code = symmetric_code(first_red)
        kill_numbers4 = [sym_code]
        if 1 <= sym_code + 1 <= 33:
            kill_numbers4.append(sym_code + 1)
        if 1 <= sym_code - 1 <= 33:
            kill_numbers4.append(sym_code - 1)
        
        results["方法4_杀号"].append(kill_numbers4)
        results["方法4_结果"].append(all(num not in current_red_balls for num in kill_numbers4))
        
        # 方法5: 红号第一位+第六位杀号法
        first_plus_last = prev_red_balls[0] + prev_red_balls[5]
        kill_numbers5 = []
        if 1 <= first_plus_last <= 33:
            kill_numbers5.append(first_plus_last)
        # 也杀掉和的一半（向下取整）
        half_sum = first_plus_last // 2
        if 1 <= half_sum <= 33:
            kill_numbers5.append(half_sum)
        
        results["方法5_杀号"].append(kill_numbers5)
        results["方法5_结果"].append(all(num not in current_red_balls for num in kill_numbers5))
        
        # 方法6: 出球顺序第一位+第二位+第五位杀号法
        # 注意：这里我们使用的是红球的顺序，不是出球顺序
        sum_125 = prev_red_balls[0] + prev_red_balls[1] + prev_red_balls[4]
        kill_numbers6 = []
        # 如果和大于33，则杀掉和除以2的结果（向下取整）
        if sum_125 > 33:
            num = sum_125 // 2
            if 1 <= num <= 33:
                kill_numbers6.append(num)
        else:
            # 否则直接杀掉和
            kill_numbers6.append(sum_125)
        
        results["方法6_杀号"].append(kill_numbers6)
        results["方法6_结果"].append(all(num not in current_red_balls for num in kill_numbers6))
        
        # 方法7: 红号之和的和杀号法
        red_sum = sum(prev_red_balls)
        digit_sum = sum_of_digits(red_sum)
        kill_numbers7 = []
        # 杀掉数字之和及其倍数
        for multiple in range(1, 4):
            num = digit_sum * multiple
            if 1 <= num <= 33:
                kill_numbers7.append(num)
        
        results["方法7_杀号"].append(kill_numbers7)
        results["方法7_结果"].append(all(num not in current_red_balls for num in kill_numbers7))
        
        # 方法8: AC值杀号法
        ac_value = calculate_ac_value(prev_red_balls)
        kill_numbers8 = []
        # 根据AC值杀号，AC值在特定范围内的号码出现概率低
        if 1 <= ac_value <= 2:
            kill_numbers8.extend([1, 4, 7, 10, 13, 16, 19, 22, 25, 28, 31])
        elif 3 <= ac_value <= 4:
            kill_numbers8.extend([2, 5, 8, 11, 14, 17, 20, 23, 26, 29, 32])
        else:  # AC值 >= 5
            kill_numbers8.extend([3, 6, 9, 12, 15, 18, 21, 24, 27, 30, 33])
        
        # 只保留1-33范围内的号码
        kill_numbers8 = [num for num in kill_numbers8 if 1 <= num <= 33]
        results["方法8_杀号"].append(kill_numbers8)
        results["方法8_结果"].append(all(num not in current_red_balls for num in kill_numbers8))
        
        # 方法9: 蓝号+AC值杀号法
        blue_ac_sum = prev_blue + ac_value
        kill_numbers9 = []
        if blue_ac_sum <= 10:
            kill_numbers9.extend([1, 5, 9, 13, 17, 21, 25, 29, 33])
        elif 11 <= blue_ac_sum <= 20:
            kill_numbers9.extend([2, 6, 10, 14, 18, 22, 26, 30])
        else:  # blue_ac_sum > 20
            kill_numbers9.extend([3, 7, 11, 15, 19, 23, 27, 31])
        
        results["方法9_杀号"].append(kill_numbers9)
        results["方法9_结果"].append(all(num not in current_red_balls for num in kill_numbers9))
        
        # 方法10: 蓝号+质数个数杀号法
        prime_count = count_primes(prev_red_balls)
        blue_prime_sum = prev_blue + prime_count
        kill_numbers10 = []
        if blue_prime_sum % 3 == 0:
            kill_numbers10.extend([3, 6, 9, 12, 15, 18, 21, 24, 27, 30, 33])
        elif blue_prime_sum % 3 == 1:
            kill_numbers10.extend([1, 4, 7, 10, 13, 16, 19, 22, 25, 28, 31])
        else:  # blue_prime_sum % 3 == 2
            kill_numbers10.extend([2, 5, 8, 11, 14, 17, 20, 23, 26, 29, 32])
        
        results["方法10_杀号"].append(kill_numbers10)
        results["方法10_结果"].append(all(num not in current_red_balls for num in kill_numbers10))
        
        # 方法11: AC值×质数个数杀号法
        ac_prime_product = ac_value * prime_count
        kill_numbers11 = []
        if 1 <= ac_prime_product <= 33:
            kill_numbers11.append(ac_prime_product)
        # 也杀掉乘积的倍数（不超过33）
        for multiple in range(2, 4):
            num = ac_prime_product * multiple
            if 1 <= num <= 33:
                kill_numbers11.append(num)
        
        results["方法11_杀号"].append(kill_numbers11)
        results["方法11_结果"].append(all(num not in current_red_balls for num in kill_numbers11))
        
        # 方法12: AC值+质数个数杀号法
        ac_prime_sum = ac_value + prime_count
        kill_numbers12 = []
        if 1 <= ac_prime_sum <= 33:
            kill_numbers12.append(ac_prime_sum)
        # 也杀掉和的倍数（不超过33）
        for multiple in range(2, 4):
            num = ac_prime_sum * multiple
            if 1 <= num <= 33:
                kill_numbers12.append(num)
        
        results["方法12_杀号"].append(kill_numbers12)
        results["方法12_结果"].append(all(num not in current_red_balls for num in kill_numbers12))
        
        # 方法13: AC值-红号第六位杀号法
        sixth_red = prev_red_balls[5]
        ac_sixth_diff = abs(ac_value - sixth_red)
        kill_numbers13 = []
        if 1 <= ac_sixth_diff <= 33:
            kill_numbers13.append(ac_sixth_diff)
        
        results["方法13_杀号"].append(kill_numbers13)
        results["方法13_结果"].append(all(num not in current_red_balls for num in kill_numbers13))
        
        # 方法14: 蓝号-红号第一、六位杀号法
        first_red = prev_red_balls[0]
        blue_first_diff = abs(prev_blue - first_red)
        blue_sixth_diff = abs(prev_blue - sixth_red)
        kill_numbers14 = []
        if 1 <= blue_first_diff <= 33:
            kill_numbers14.append(blue_first_diff)
        if 1 <= blue_sixth_diff <= 33 and blue_sixth_diff != blue_first_diff:
            kill_numbers14.append(blue_sixth_diff)
        
        results["方法14_杀号"].append(kill_numbers14)
        results["方法14_结果"].append(all(num not in current_red_balls for num in kill_numbers14))
        
        # 方法15: 蓝号+AC值+质数个数杀号法
        blue_ac_prime_sum = prev_blue + ac_value + prime_count
        kill_numbers15 = []
        if 1 <= blue_ac_prime_sum <= 33:
            kill_numbers15.append(blue_ac_prime_sum)
        else:
            # 如果和大于33，则取余数
            remainder = blue_ac_prime_sum % 33
            if remainder > 0:  # 确保余数不为0
                kill_numbers15.append(remainder)
            else:
                kill_numbers15.append(33)  # 如果余数为0，则杀号33
        
        results["方法15_杀号"].append(kill_numbers15)
        results["方法15_结果"].append(all(num not in current_red_balls for num in kill_numbers15))
        
        # 方法16: 红号×0.88杀号法
        kill_numbers16 = []
        for red in prev_red_balls:
            num = int(red * 0.88)
            if 1 <= num <= 33 and num not in kill_numbers16:
                kill_numbers16.append(num)
        
        results["方法16_杀号"].append(kill_numbers16)
        results["方法16_结果"].append(all(num not in current_red_balls for num in kill_numbers16))
        
        # 方法17: 红号第三位对称码+7杀号法
        third_red = prev_red_balls[2]
        third_sym_plus7 = symmetric_code(third_red) + 7
        kill_numbers17 = []
        if 1 <= third_sym_plus7 <= 33:
            kill_numbers17.append(third_sym_plus7)
        
        results["方法17_杀号"].append(kill_numbers17)
        results["方法17_结果"].append(all(num not in current_red_balls for num in kill_numbers17))
    
    # 计算各方法的成功率
    success_rates = {}
    for method in range(1, 18):
        method_key = f"方法{method}_结果"
        valid_results = [r for r in results[method_key] if r is not None]
        if valid_results:
            success_count = sum(valid_results)
            total_count = len(valid_results)
            success_rate = success_count / total_count
            success_rates[f"方法{method}"] = success_rate
    
    # 生成最新一期的杀号预测
    if len(periods) >= 1:
        latest_idx = len(periods) - 1
        latest_red_balls = []
        for j in range(1, 7):
            col_name = f"红球号码{j}"
            latest_red_balls.append(df.iloc[latest_idx][col_name])
        
        latest_blue = blue_balls[latest_idx]
        
        # 方法1: 蓝号杀号法
        next_kill1 = []
        for multiple in range(1, 4):
            num = latest_blue * multiple
            if 1 <= num <= 33:
                next_kill1.append(num)
        
        # 方法2: 红号尾数之和杀号法
        tail_sum = sum_of_tails(latest_red_balls)
        next_kill2 = []
        for multiple in range(1, 4):
            num = tail_sum * multiple
            if 1 <= num <= 33:
                next_kill2.append(num)
        
        # 方法3: 质数个数杀号法
        prime_count = count_primes(latest_red_balls)
        next_kill3 = []
        for multiple in range(1, 6):
            num = prime_count * multiple
            if 1 <= num <= 33:
                next_kill3.append(num)
        
        # 方法4: 红号第一位对称码杀号法
        first_red = latest_red_balls[0]
        sym_code = symmetric_code(first_red)
        next_kill4 = [sym_code]
        if 1 <= sym_code + 1 <= 33:
            next_kill4.append(sym_code + 1)
        if 1 <= sym_code - 1 <= 33:
            next_kill4.append(sym_code - 1)
        
        # 方法5: 红号第一位+第六位杀号法
        first_plus_last = latest_red_balls[0] + latest_red_balls[5]
        next_kill5 = []
        if 1 <= first_plus_last <= 33:
            next_kill5.append(first_plus_last)
        half_sum = first_plus_last // 2
        if 1 <= half_sum <= 33:
            next_kill5.append(half_sum)
        
        # 方法6: 出球顺序第一位+第二位+第五位杀号法
        sum_125 = latest_red_balls[0] + latest_red_balls[1] + latest_red_balls[4]
        next_kill6 = []
        if sum_125 > 33:
            num = sum_125 // 2
            if 1 <= num <= 33:
                next_kill6.append(num)
        else:
            next_kill6.append(sum_125)
        
        # 方法7: 红号之和的和杀号法
        red_sum = sum(latest_red_balls)
        digit_sum = sum_of_digits(red_sum)
        next_kill7 = []
        # 杀掉数字之和及其倍数
        for multiple in range(1, 4):
            num = digit_sum * multiple
            if 1 <= num <= 33:
                next_kill7.append(num)
        
        # 方法8: AC值杀号法
        ac_value = calculate_ac_value(latest_red_balls)
        next_kill8 = []
        # 根据AC值杀号，AC值在特定范围内的号码出现概率低
        if 1 <= ac_value <= 2:
            next_kill8.extend([1, 4, 7, 10, 13, 16, 19, 22, 25, 28, 31])
        elif 3 <= ac_value <= 4:
            next_kill8.extend([2, 5, 8, 11, 14, 17, 20, 23, 26, 29, 32])
        else:  # AC值 >= 5
            next_kill8.extend([3, 6, 9, 12, 15, 18, 21, 24, 27, 30, 33])
        
        # 只保留1-33范围内的号码
        next_kill8 = [num for num in next_kill8 if 1 <= num <= 33]
        
        # 方法9: 蓝号+AC值杀号法
        blue_ac_sum = latest_blue + ac_value
        next_kill9 = []
        if blue_ac_sum <= 10:
            next_kill9.extend([1, 5, 9, 13, 17, 21, 25, 29, 33])
        elif 11 <= blue_ac_sum <= 20:
            next_kill9.extend([2, 6, 10, 14, 18, 22, 26, 30])
        else:  # blue_ac_sum > 20
            next_kill9.extend([3, 7, 11, 15, 19, 23, 27, 31])
        
        # 方法10: 蓝号+质数个数杀号法
        prime_count = count_primes(latest_red_balls)
        blue_prime_sum = latest_blue + prime_count
        next_kill10 = []
        if blue_prime_sum % 3 == 0:
            next_kill10.extend([3, 6, 9, 12, 15, 18, 21, 24, 27, 30, 33])
        elif blue_prime_sum % 3 == 1:
            next_kill10.extend([1, 4, 7, 10, 13, 16, 19, 22, 25, 28, 31])
        else:  # blue_prime_sum % 3 == 2
            next_kill10.extend([2, 5, 8, 11, 14, 17, 20, 23, 26, 29, 32])
        
        # 方法11: AC值×质数个数杀号法
        ac_prime_product = ac_value * prime_count
        next_kill11 = []
        if 1 <= ac_prime_product <= 33:
            next_kill11.append(ac_prime_product)
        # 也杀掉乘积的倍数（不超过33）
        for multiple in range(2, 4):
            num = ac_prime_product * multiple
            if 1 <= num <= 33:
                next_kill11.append(num)
        
        # 方法12: AC值+质数个数杀号法
        ac_prime_sum = ac_value + prime_count
        next_kill12 = []
        if 1 <= ac_prime_sum <= 33:
            next_kill12.append(ac_prime_sum)
        # 也杀掉和的倍数（不超过33）
        for multiple in range(2, 4):
            num = ac_prime_sum * multiple
            if 1 <= num <= 33:
                next_kill12.append(num)
        
        # 方法13: AC值-红号第六位杀号法
        sixth_red = latest_red_balls[5]
        ac_sixth_diff = abs(ac_value - sixth_red)
        next_kill13 = []
        if 1 <= ac_sixth_diff <= 33:
            next_kill13.append(ac_sixth_diff)
        
        # 方法14: 蓝号-红号第一、六位杀号法
        first_red = latest_red_balls[0]
        blue_first_diff = abs(latest_blue - first_red)
        blue_sixth_diff = abs(latest_blue - sixth_red)
        next_kill14 = []
        if 1 <= blue_first_diff <= 33:
            next_kill14.append(blue_first_diff)
        if 1 <= blue_sixth_diff <= 33 and blue_sixth_diff != blue_first_diff:
            next_kill14.append(blue_sixth_diff)
        
        # 方法15: 蓝号+AC值+质数个数杀号法
        blue_ac_prime_sum = latest_blue + ac_value + prime_count
        next_kill15 = []
        if 1 <= blue_ac_prime_sum <= 33:
            next_kill15.append(blue_ac_prime_sum)
        else:
            # 如果和大于33，则取余数
            remainder = blue_ac_prime_sum % 33
            if remainder > 0:  # 确保余数不为0
                next_kill15.append(remainder)
            else:
                next_kill15.append(33)  # 如果余数为0，则杀号33
        
        # 方法16: 红号×0.88杀号法
        next_kill16 = []
        for red in latest_red_balls:
            num = int(red * 0.88)
            if 1 <= num <= 33 and num not in next_kill16:
                next_kill16.append(num)
        
        # 方法17: 红号第三位对称码+7杀号法
        third_red = latest_red_balls[2]
        third_sym_plus7 = symmetric_code(third_red) + 7
        next_kill17 = []
        if 1 <= third_sym_plus7 <= 33:
            next_kill17.append(third_sym_plus7)
        
        next_predictions = {
            "方法1": next_kill1,
            "方法2": next_kill2,
            "方法3": next_kill3,
            "方法4": next_kill4,
            "方法5": next_kill5,
            "方法6": next_kill6,
            "方法7": next_kill7,
            "方法8": next_kill8,
            "方法9": next_kill9,
            "方法10": next_kill10,
            "方法11": next_kill11,
            "方法12": next_kill12,
            "方法13": next_kill13,
            "方法14": next_kill14,
            "方法15": next_kill15,
            "方法16": next_kill16,
            "方法17": next_kill17
        }
    else:
        next_predictions = {}
    
    # 创建结果DataFrame
    results_df = pd.DataFrame({
        "期数": results["期数"],
        "方法1_杀号": [','.join(map(str, x)) for x in results["方法1_杀号"]],
        "方法1_结果": results["方法1_结果"],
        "方法2_杀号": [','.join(map(str, x)) for x in results["方法2_杀号"]],
        "方法2_结果": results["方法2_结果"],
        "方法3_杀号": [','.join(map(str, x)) for x in results["方法3_杀号"]],
        "方法3_结果": results["方法3_结果"],
        "方法4_杀号": [','.join(map(str, x)) for x in results["方法4_杀号"]],
        "方法4_结果": results["方法4_结果"],
        "方法5_杀号": [','.join(map(str, x)) for x in results["方法5_杀号"]],
        "方法5_结果": results["方法5_结果"],
        "方法6_杀号": [','.join(map(str, x)) for x in results["方法6_杀号"]],
        "方法6_结果": results["方法6_结果"],
        "方法7_杀号": [','.join(map(str, x)) for x in results["方法7_杀号"]],
        "方法7_结果": results["方法7_结果"],
        "方法8_杀号": [','.join(map(str, x)) for x in results["方法8_杀号"]],
        "方法8_结果": results["方法8_结果"],
        "方法9_杀号": [','.join(map(str, x)) for x in results["方法9_杀号"]],
        "方法9_结果": results["方法9_结果"],
        "方法10_杀号": [','.join(map(str, x)) for x in results["方法10_杀号"]],
        "方法10_结果": results["方法10_结果"],
        "方法11_杀号": [','.join(map(str, x)) for x in results["方法11_杀号"]],
        "方法11_结果": results["方法11_结果"],
        "方法12_杀号": [','.join(map(str, x)) for x in results["方法12_杀号"]],
        "方法12_结果": results["方法12_结果"],
        "方法13_杀号": [','.join(map(str, x)) for x in results["方法13_杀号"]],
        "方法13_结果": results["方法13_结果"],
        "方法14_杀号": [','.join(map(str, x)) for x in results["方法14_杀号"]],
        "方法14_结果": results["方法14_结果"],
        "方法15_杀号": [','.join(map(str, x)) for x in results["方法15_杀号"]],
        "方法15_结果": results["方法15_结果"],
        "方法16_杀号": [','.join(map(str, x)) for x in results["方法16_杀号"]],
        "方法16_结果": results["方法16_结果"],
        "方法17_杀号": [','.join(map(str, x)) for x in results["方法17_杀号"]],
        "方法17_结果": results["方法17_结果"],
    })
    
    return results_df, success_rates, next_predictions


# def write_red_kill_results_to_excel(excel_file, results_df, success_rates, next_predictions):
#     """
#     将红球杀号结果写入Excel文件的新工作表，并增加4列：
#       1. 成功杀号号码：本期所有成功杀号的号码（取并集）
#       2. 成功杀号个数：成功杀号号码的个数
#       3. 剩余可用号码：1~33中不在成功杀号号码内的号码
#       4. 剩余可用号码(包括失败)：1~33中不在所有方法所给出的杀号号码内的号码（不论成功或失败）
#
#     同时在写入下一期预测时，统计所有方法预测中杀掉的号码的并集，
#     并计算1~33中未被杀掉的号码作为剩余可用号码，写入到表中。
#     """
#     from openpyxl import load_workbook
#     from openpyxl.styles import PatternFill
#     import logging
#
#     try:
#         # 加载工作簿
#         workbook = load_workbook(excel_file)
#
#         # 检查是否已存在红球杀号工作表，如果存在则删除
#         if "红球杀号" in workbook.sheetnames:
#             del workbook["红球杀号"]
#
#         # 创建新的工作表
#         sheet = workbook.create_sheet("红球杀号")
#
#         # 定义填充颜色
#         success_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")  # 绿色
#         fail_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # 红色
#
#         # 写入标题行（增加4列）
#         headers = ["期数",
#                    "方法1杀号", "方法1结果",
#                    "方法2杀号", "方法2结果",
#                    "方法3杀号", "方法3结果",
#                    "方法4杀号", "方法4结果",
#                    "方法5杀号", "方法5结果",
#                    "方法6杀号", "方法6结果",
#                    "方法7杀号", "方法7结果",
#                    "方法8杀号", "方法8结果",
#                    "方法9杀号", "方法9结果",
#                    "方法10杀号", "方法10结果",
#                    "方法11杀号", "方法11结果",
#                    "方法12杀号", "方法12结果",
#                    "方法13杀号", "方法13结果",
#                    "方法14杀号", "方法14结果",
#                    "方法15杀号", "方法15结果",
#                    "方法16杀号", "方法16结果",
#                    "方法17杀号", "方法17结果",
#                    "成功杀号号码", "成功杀号个数", "剩余可用号码", "剩余可用号码(包括失败)"]
#         for col, header in enumerate(headers, 1):
#             sheet.cell(row=1, column=col).value = header
#         total_methods = 17  # 方法数固定为17
#
#         # 写入数据部分
#         for row_idx, row in results_df.iterrows():
#             # 写入前35列数据（期数及各方法数据）
#             sheet.cell(row=row_idx + 2, column=1).value = row["期数"]
#
#             sheet.cell(row=row_idx + 2, column=2).value = row["方法1_杀号"]
#             cell = sheet.cell(row=row_idx + 2, column=3)
#             cell.value = "成功" if row["方法1_结果"] else "失败" if row["方法1_结果"] is not None else ""
#             if row["方法1_结果"] is not None:
#                 cell.fill = success_fill if row["方法1_结果"] else fail_fill
#
#             sheet.cell(row=row_idx + 2, column=4).value = row["方法2_杀号"]
#             cell = sheet.cell(row=row_idx + 2, column=5)
#             cell.value = "成功" if row["方法2_结果"] else "失败" if row["方法2_结果"] is not None else ""
#             if row["方法2_结果"] is not None:
#                 cell.fill = success_fill if row["方法2_结果"] else fail_fill
#
#             sheet.cell(row=row_idx + 2, column=6).value = row["方法3_杀号"]
#             cell = sheet.cell(row=row_idx + 2, column=7)
#             cell.value = "成功" if row["方法3_结果"] else "失败" if row["方法3_结果"] is not None else ""
#             if row["方法3_结果"] is not None:
#                 cell.fill = success_fill if row["方法3_结果"] else fail_fill
#
#             sheet.cell(row=row_idx + 2, column=8).value = row["方法4_杀号"]
#             cell = sheet.cell(row=row_idx + 2, column=9)
#             cell.value = "成功" if row["方法4_结果"] else "失败" if row["方法4_结果"] is not None else ""
#             if row["方法4_结果"] is not None:
#                 cell.fill = success_fill if row["方法4_结果"] else fail_fill
#
#             sheet.cell(row=row_idx + 2, column=10).value = row["方法5_杀号"]
#             cell = sheet.cell(row=row_idx + 2, column=11)
#             cell.value = "成功" if row["方法5_结果"] else "失败" if row["方法5_结果"] is not None else ""
#             if row["方法5_结果"] is not None:
#                 cell.fill = success_fill if row["方法5_结果"] else fail_fill
#
#             sheet.cell(row=row_idx + 2, column=12).value = row["方法6_杀号"]
#             cell = sheet.cell(row=row_idx + 2, column=13)
#             cell.value = "成功" if row["方法6_结果"] else "失败" if row["方法6_结果"] is not None else ""
#             if row["方法6_结果"] is not None:
#                 cell.fill = success_fill if row["方法6_结果"] else fail_fill
#
#             sheet.cell(row=row_idx + 2, column=14).value = row["方法7_杀号"]
#             cell = sheet.cell(row=row_idx + 2, column=15)
#             cell.value = "成功" if row["方法7_结果"] else "失败" if row["方法7_结果"] is not None else ""
#             if row["方法7_结果"] is not None:
#                 cell.fill = success_fill if row["方法7_结果"] else fail_fill
#
#             sheet.cell(row=row_idx + 2, column=16).value = row["方法8_杀号"]
#             cell = sheet.cell(row=row_idx + 2, column=17)
#             cell.value = "成功" if row["方法8_结果"] else "失败" if row["方法8_结果"] is not None else ""
#             if row["方法8_结果"] is not None:
#                 cell.fill = success_fill if row["方法8_结果"] else fail_fill
#
#             sheet.cell(row=row_idx + 2, column=18).value = row["方法9_杀号"]
#             cell = sheet.cell(row=row_idx + 2, column=19)
#             cell.value = "成功" if row["方法9_结果"] else "失败" if row["方法9_结果"] is not None else ""
#             if row["方法9_结果"] is not None:
#                 cell.fill = success_fill if row["方法9_结果"] else fail_fill
#
#             sheet.cell(row=row_idx + 2, column=20).value = row["方法10_杀号"]
#             cell = sheet.cell(row=row_idx + 2, column=21)
#             cell.value = "成功" if row["方法10_结果"] else "失败" if row["方法10_结果"] is not None else ""
#             if row["方法10_结果"] is not None:
#                 cell.fill = success_fill if row["方法10_结果"] else fail_fill
#
#             sheet.cell(row=row_idx + 2, column=22).value = row["方法11_杀号"]
#             cell = sheet.cell(row=row_idx + 2, column=23)
#             cell.value = "成功" if row["方法11_结果"] else "失败" if row["方法11_结果"] is not None else ""
#             if row["方法11_结果"] is not None:
#                 cell.fill = success_fill if row["方法11_结果"] else fail_fill
#
#             sheet.cell(row=row_idx + 2, column=24).value = row["方法12_杀号"]
#             cell = sheet.cell(row=row_idx + 2, column=25)
#             cell.value = "成功" if row["方法12_结果"] else "失败" if row["方法12_结果"] is not None else ""
#             if row["方法12_结果"] is not None:
#                 cell.fill = success_fill if row["方法12_结果"] else fail_fill
#
#             sheet.cell(row=row_idx + 2, column=26).value = row["方法13_杀号"]
#             cell = sheet.cell(row=row_idx + 2, column=27)
#             cell.value = "成功" if row["方法13_结果"] else "失败" if row["方法13_结果"] is not None else ""
#             if row["方法13_结果"] is not None:
#                 cell.fill = success_fill if row["方法13_结果"] else fail_fill
#
#             sheet.cell(row=row_idx + 2, column=28).value = row["方法14_杀号"]
#             cell = sheet.cell(row=row_idx + 2, column=29)
#             cell.value = "成功" if row["方法14_结果"] else "失败" if row["方法14_结果"] is not None else ""
#             if row["方法14_结果"] is not None:
#                 cell.fill = success_fill if row["方法14_结果"] else fail_fill
#
#             sheet.cell(row=row_idx + 2, column=30).value = row["方法15_杀号"]
#             cell = sheet.cell(row=row_idx + 2, column=31)
#             cell.value = "成功" if row["方法15_结果"] else "失败" if row["方法15_结果"] is not None else ""
#             if row["方法15_结果"] is not None:
#                 cell.fill = success_fill if row["方法15_结果"] else fail_fill
#
#             sheet.cell(row=row_idx + 2, column=32).value = row["方法16_杀号"]
#             cell = sheet.cell(row=row_idx + 2, column=33)
#             cell.value = "成功" if row["方法16_结果"] else "失败" if row["方法16_结果"] is not None else ""
#             if row["方法16_结果"] is not None:
#                 cell.fill = success_fill if row["方法16_结果"] else fail_fill
#
#             sheet.cell(row=row_idx + 2, column=34).value = row["方法17_杀号"]
#             cell = sheet.cell(row=row_idx + 2, column=35)
#             cell.value = "成功" if row["方法17_结果"] else "失败" if row["方法17_结果"] is not None else ""
#             if row["方法17_结果"] is not None:
#                 cell.fill = success_fill if row["方法17_结果"] else fail_fill
#
#             # 计算新增4列数据：成功杀号号码、成功杀号个数、剩余可用号码（成功杀号）、剩余可用号码（所有方法）
#             success_numbers_set = set()
#             all_killed_numbers_set = set()
#             for i in range(1, 18):
#                 result_field = f"方法{i}_结果"
#                 kill_field = f"方法{i}_杀号"
#                 kill_value = row[kill_field]
#                 if kill_value is not None:
#                     if isinstance(kill_value, str):
#                         nums = [int(x.strip()) for x in kill_value.split(',') if x.strip().isdigit()]
#                     elif isinstance(kill_value, (list, tuple)):
#                         nums = kill_value
#                     else:
#                         nums = [kill_value]
#                     all_killed_numbers_set.update(nums)
#                 if row[result_field]:
#                     if kill_value is not None:
#                         if isinstance(kill_value, str):
#                             nums = [int(x.strip()) for x in kill_value.split(',') if x.strip().isdigit()]
#                         elif isinstance(kill_value, (list, tuple)):
#                             nums = kill_value
#                         else:
#                             nums = [kill_value]
#                         success_numbers_set.update(nums)
#             success_numbers = sorted(success_numbers_set)
#             success_count = len(success_numbers)
#             all_numbers = set(range(1, 34))
#             remaining_numbers_success = sorted(all_numbers - success_numbers_set)
#             remaining_numbers_all = sorted(all_numbers - all_killed_numbers_set)
#
#             sheet.cell(row=row_idx + 2, column=36).value = ','.join(map(str, success_numbers))
#             sheet.cell(row=row_idx + 2, column=37).value = success_count
#             sheet.cell(row=row_idx + 2, column=38).value = ','.join(map(str, remaining_numbers_success))
#             sheet.cell(row=row_idx + 2, column=39).value = ','.join(map(str, remaining_numbers_all))
#
#         # 统计所有期次的杀号个数以及所有方法的成功次数
#         kill_counts = []  # 每期成功杀号号码的个数
#         overall_success_count = 0  # 所有方法成功次数累计
#         all_methods_success_count = 0  # 每期所有方法均成功的次数
#         for idx, row in results_df.iterrows():
#             success_numbers_set = set()
#             all_success = True
#             for i in range(1, total_methods + 1):
#                 result_field = f"方法{i}_结果"
#                 kill_field = f"方法{i}_杀号"
#                 if row[result_field]:
#                     overall_success_count += 1
#                     kill_value = row[kill_field]
#                     if kill_value is not None:
#                         if isinstance(kill_value, str):
#                             nums = [int(x.strip()) for x in kill_value.split(',') if x.strip().isdigit()]
#                         elif isinstance(kill_value, (list, tuple)):
#                             nums = kill_value
#                         else:
#                             nums = [kill_value]
#                         success_numbers_set.update(nums)
#                 else:
#                     all_success = False
#             kill_counts.append(len(success_numbers_set))
#             if all_success:
#                 all_methods_success_count += 1
#
#         max_kill_count = max(kill_counts) if kill_counts else 0
#         min_kill_count = min(kill_counts) if kill_counts else 0
#         avg_kill_count = sum(kill_counts) / len(kill_counts) if kill_counts else 0
#         combined_success_rate = overall_success_count / (total_methods * len(results_df)) if len(results_df) > 0 else 0
#
#         # 新增统计：成功杀号个数达到某个阈值的次数
#         kill_ge_10_count = sum(1 for count in kill_counts if count >= 10)
#         kill_ge_15_count = sum(1 for count in kill_counts if count >= 15)
#         kill_ge_20_count = sum(1 for count in kill_counts if count >= 20)
#         # 写入成功率统计
#         row_offset = len(results_df) + 3
#         sheet.cell(row=row_offset, column=1).value = "杀号方法成功率统计"
#         for i, (method, rate) in enumerate(success_rates.items(), 1):
#             sheet.cell(row=row_offset + i, column=1).value = method
#             cell = sheet.cell(row=row_offset + i, column=2)
#             cell.value = f"{rate:.2%}"
#             if rate >= 0.8:
#                 cell.fill = success_fill
#             elif rate < 0.5:
#                 cell.fill = fail_fill
#
#         # 写入下一期预测部分
#         row_offset = row_offset + len(success_rates) + 2
#         next_period = results_df["期数"].iloc[-1] + 1
#         sheet.cell(row=row_offset, column=1).value = f"{next_period}期红球杀号预测"
#         # 写入预测表头：方法、杀号
#         sheet.cell(row=row_offset + 1, column=1).value = "方法"
#         sheet.cell(row=row_offset + 1, column=2).value = "杀号"
#         # 写入各方法预测结果，同时计算所有方法预测的杀号号码并集
#         union_next_killed = set()
#         for i, (method, numbers) in enumerate(next_predictions.items(), 1):
#             row_index = row_offset + 1 + i
#             sheet.cell(row=row_index, column=1).value = method
#             kill_numbers_str = ','.join(map(str, numbers))
#             sheet.cell(row=row_index, column=2).value = kill_numbers_str
#             # 更新所有方法预测的号码并集
#             if isinstance(numbers, (list, tuple)):
#                 union_next_killed.update(numbers)
#             else:
#                 union_next_killed.add(numbers)
#         # 计算剩余可用号码（1~33中未被所有方法杀掉的号码）
#         remaining_numbers_overall = sorted(set(range(1, 34)) - union_next_killed)
#         row_index = row_offset + 1 + len(next_predictions) + 1
#         sheet.cell(row=row_index, column=1).value = "剩余可用号码"
#         sheet.cell(row=row_index, column=2).value = ','.join(map(str, remaining_numbers_overall))
#
#         # 写入额外统计项目
#         stat_start = row_index + 2
#         sheet.cell(row=stat_start, column=1).value = "最多杀号个数"
#         sheet.cell(row=stat_start, column=2).value = max_kill_count
#
#         sheet.cell(row=stat_start + 1, column=1).value = "最少杀号个数"
#         sheet.cell(row=stat_start + 1, column=2).value = min_kill_count
#
#         sheet.cell(row=stat_start + 2, column=1).value = "平均杀号个数"
#         sheet.cell(row=stat_start + 2, column=2).value = avg_kill_count
#
#         sheet.cell(row=stat_start + 3, column=1).value = "所有方法综合成功率"
#         cell = sheet.cell(row=stat_start + 3, column=2)
#         cell.value = f"{combined_success_rate:.2%}"
#         if combined_success_rate >= 0.8:
#             cell.fill = success_fill
#         elif combined_success_rate < 0.5:
#             cell.fill = fail_fill
#
#         sheet.cell(row=stat_start + 4, column=1).value = "所有方法都成功的次数"
#         sheet.cell(row=stat_start + 4, column=2).value = all_methods_success_count
#
#         sheet.cell(row=stat_start + 5, column=1).value = "成功杀号10个以上的次数"
#         sheet.cell(row=stat_start + 5, column=2).value = kill_ge_10_count
#
#         sheet.cell(row=stat_start + 6, column=1).value = "成功杀号15个以上的次数"
#         sheet.cell(row=stat_start + 6, column=2).value = kill_ge_15_count
#
#         sheet.cell(row=stat_start + 7, column=1).value = "成功杀号20个以上的次数"
#         sheet.cell(row=stat_start + 7, column=2).value = kill_ge_20_count
#
#         # 保存工作簿
#         workbook.save(excel_file)
#         logging.info(f"红球杀号结果已写入Excel文件 '{excel_file}' 的 '红球杀号' 工作表")
#         return True
#     except Exception as e:
#         logging.error(f"写入Excel时发生错误: {e}")
#         return False

def write_red_kill_results_to_excel(excel_file, results_df, success_rates, next_predictions):
    """
    将红球杀号结果写入Excel文件的新工作表，并增加4列：
      1. 成功杀号号码：本期所有成功杀号的号码（取并集）
      2. 成功杀号个数：成功杀号号码的个数
      3. 剩余可用号码：1~33中不在成功杀号号码内的号码
      4. 剩余可用号码(包括失败)：1~33中不在所有方法所给出的杀号号码内的号码（不论成功或失败）

    同时在写入下一期预测时，统计所有方法预测中杀掉的号码的并集，
    并计算1~33中未被杀掉的号码作为剩余可用号码。但这里修改为：根据各方法成功率，
    给1～33的号码计算风险分数，并将号码分为不同优先级（风险越低优先级越高），写入到表中。
    """
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill
    import logging

    try:
        # 加载工作簿
        workbook = load_workbook(excel_file)

        # 检查是否已存在红球杀号工作表，如果存在则删除
        if "红球杀号" in workbook.sheetnames:
            del workbook["红球杀号"]

        # 创建新的工作表
        sheet = workbook.create_sheet("红球杀号")

        # 定义填充颜色
        success_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")  # 绿色
        fail_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # 红色

        # 写入标题行（增加4列）
        headers = ["期数",
                   "方法1杀号", "方法1结果",
                   "方法2杀号", "方法2结果",
                   "方法3杀号", "方法3结果",
                   "方法4杀号", "方法4结果",
                   "方法5杀号", "方法5结果",
                   "方法6杀号", "方法6结果",
                   "方法7杀号", "方法7结果",
                   "方法8杀号", "方法8结果",
                   "方法9杀号", "方法9结果",
                   "方法10杀号", "方法10结果",
                   "方法11杀号", "方法11结果",
                   "方法12杀号", "方法12结果",
                   "方法13杀号", "方法13结果",
                   "方法14杀号", "方法14结果",
                   "方法15杀号", "方法15结果",
                   "方法16杀号", "方法16结果",
                   "方法17杀号", "方法17结果",
                   "成功杀号号码", "成功杀号个数", "剩余可用号码", "剩余可用号码(包括失败)"]
        for col, header in enumerate(headers, 1):
            sheet.cell(row=1, column=col).value = header
        total_methods = 17  # 方法数固定为17

        # 写入数据部分
        for row_idx, row in results_df.iterrows():
            # 写入前35列数据（期数及各方法数据）
            sheet.cell(row=row_idx + 2, column=1).value = row["期数"]

            sheet.cell(row=row_idx + 2, column=2).value = row["方法1_杀号"]
            cell = sheet.cell(row=row_idx + 2, column=3)
            cell.value = "成功" if row["方法1_结果"] else "失败" if row["方法1_结果"] is not None else ""
            if row["方法1_结果"] is not None:
                cell.fill = success_fill if row["方法1_结果"] else fail_fill

            sheet.cell(row=row_idx + 2, column=4).value = row["方法2_杀号"]
            cell = sheet.cell(row=row_idx + 2, column=5)
            cell.value = "成功" if row["方法2_结果"] else "失败" if row["方法2_结果"] is not None else ""
            if row["方法2_结果"] is not None:
                cell.fill = success_fill if row["方法2_结果"] else fail_fill

            sheet.cell(row=row_idx + 2, column=6).value = row["方法3_杀号"]
            cell = sheet.cell(row=row_idx + 2, column=7)
            cell.value = "成功" if row["方法3_结果"] else "失败" if row["方法3_结果"] is not None else ""
            if row["方法3_结果"] is not None:
                cell.fill = success_fill if row["方法3_结果"] else fail_fill

            sheet.cell(row=row_idx + 2, column=8).value = row["方法4_杀号"]
            cell = sheet.cell(row=row_idx + 2, column=9)
            cell.value = "成功" if row["方法4_结果"] else "失败" if row["方法4_结果"] is not None else ""
            if row["方法4_结果"] is not None:
                cell.fill = success_fill if row["方法4_结果"] else fail_fill

            sheet.cell(row=row_idx + 2, column=10).value = row["方法5_杀号"]
            cell = sheet.cell(row=row_idx + 2, column=11)
            cell.value = "成功" if row["方法5_结果"] else "失败" if row["方法5_结果"] is not None else ""
            if row["方法5_结果"] is not None:
                cell.fill = success_fill if row["方法5_结果"] else fail_fill

            sheet.cell(row=row_idx + 2, column=12).value = row["方法6_杀号"]
            cell = sheet.cell(row=row_idx + 2, column=13)
            cell.value = "成功" if row["方法6_结果"] else "失败" if row["方法6_结果"] is not None else ""
            if row["方法6_结果"] is not None:
                cell.fill = success_fill if row["方法6_结果"] else fail_fill

            sheet.cell(row=row_idx + 2, column=14).value = row["方法7_杀号"]
            cell = sheet.cell(row=row_idx + 2, column=15)
            cell.value = "成功" if row["方法7_结果"] else "失败" if row["方法7_结果"] is not None else ""
            if row["方法7_结果"] is not None:
                cell.fill = success_fill if row["方法7_结果"] else fail_fill

            sheet.cell(row=row_idx + 2, column=16).value = row["方法8_杀号"]
            cell = sheet.cell(row=row_idx + 2, column=17)
            cell.value = "成功" if row["方法8_结果"] else "失败" if row["方法8_结果"] is not None else ""
            if row["方法8_结果"] is not None:
                cell.fill = success_fill if row["方法8_结果"] else fail_fill

            sheet.cell(row=row_idx + 2, column=18).value = row["方法9_杀号"]
            cell = sheet.cell(row=row_idx + 2, column=19)
            cell.value = "成功" if row["方法9_结果"] else "失败" if row["方法9_结果"] is not None else ""
            if row["方法9_结果"] is not None:
                cell.fill = success_fill if row["方法9_结果"] else fail_fill

            sheet.cell(row=row_idx + 2, column=20).value = row["方法10_杀号"]
            cell = sheet.cell(row=row_idx + 2, column=21)
            cell.value = "成功" if row["方法10_结果"] else "失败" if row["方法10_结果"] is not None else ""
            if row["方法10_结果"] is not None:
                cell.fill = success_fill if row["方法10_结果"] else fail_fill

            sheet.cell(row=row_idx + 2, column=22).value = row["方法11_杀号"]
            cell = sheet.cell(row=row_idx + 2, column=23)
            cell.value = "成功" if row["方法11_结果"] else "失败" if row["方法11_结果"] is not None else ""
            if row["方法11_结果"] is not None:
                cell.fill = success_fill if row["方法11_结果"] else fail_fill

            sheet.cell(row=row_idx + 2, column=24).value = row["方法12_杀号"]
            cell = sheet.cell(row=row_idx + 2, column=25)
            cell.value = "成功" if row["方法12_结果"] else "失败" if row["方法12_结果"] is not None else ""
            if row["方法12_结果"] is not None:
                cell.fill = success_fill if row["方法12_结果"] else fail_fill

            sheet.cell(row=row_idx + 2, column=26).value = row["方法13_杀号"]
            cell = sheet.cell(row=row_idx + 2, column=27)
            cell.value = "成功" if row["方法13_结果"] else "失败" if row["方法13_结果"] is not None else ""
            if row["方法13_结果"] is not None:
                cell.fill = success_fill if row["方法13_结果"] else fail_fill

            sheet.cell(row=row_idx + 2, column=28).value = row["方法14_杀号"]
            cell = sheet.cell(row=row_idx + 2, column=29)
            cell.value = "成功" if row["方法14_结果"] else "失败" if row["方法14_结果"] is not None else ""
            if row["方法14_结果"] is not None:
                cell.fill = success_fill if row["方法14_结果"] else fail_fill

            sheet.cell(row=row_idx + 2, column=30).value = row["方法15_杀号"]
            cell = sheet.cell(row=row_idx + 2, column=31)
            cell.value = "成功" if row["方法15_结果"] else "失败" if row["方法15_结果"] is not None else ""
            if row["方法15_结果"] is not None:
                cell.fill = success_fill if row["方法15_结果"] else fail_fill

            sheet.cell(row=row_idx + 2, column=32).value = row["方法16_杀号"]
            cell = sheet.cell(row=row_idx + 2, column=33)
            cell.value = "成功" if row["方法16_结果"] else "失败" if row["方法16_结果"] is not None else ""
            if row["方法16_结果"] is not None:
                cell.fill = success_fill if row["方法16_结果"] else fail_fill

            sheet.cell(row=row_idx + 2, column=34).value = row["方法17_杀号"]
            cell = sheet.cell(row=row_idx + 2, column=35)
            cell.value = "成功" if row["方法17_结果"] else "失败" if row["方法17_结果"] is not None else ""
            if row["方法17_结果"] is not None:
                cell.fill = success_fill if row["方法17_结果"] else fail_fill

            # 计算新增4列数据：成功杀号号码、成功杀号个数、剩余可用号码（成功杀号）、剩余可用号码（所有方法）
            success_numbers_set = set()
            all_killed_numbers_set = set()
            for i in range(1, 18):
                result_field = f"方法{i}_结果"
                kill_field = f"方法{i}_杀号"
                kill_value = row[kill_field]
                if kill_value is not None:
                    if isinstance(kill_value, str):
                        nums = [int(x.strip()) for x in kill_value.split(',') if x.strip().isdigit()]
                    elif isinstance(kill_value, (list, tuple)):
                        nums = kill_value
                    else:
                        nums = [kill_value]
                    all_killed_numbers_set.update(nums)
                if row[result_field]:
                    if kill_value is not None:
                        if isinstance(kill_value, str):
                            nums = [int(x.strip()) for x in kill_value.split(',') if x.strip().isdigit()]
                        elif isinstance(kill_value, (list, tuple)):
                            nums = kill_value
                        else:
                            nums = [kill_value]
                        success_numbers_set.update(nums)
            success_numbers = sorted(success_numbers_set)
            success_count = len(success_numbers)
            # ----- 新增：把每个成功号码对应的方法都收集起来 -----
            number_sources = {}  # num -> [方法i,…]
            for i in range(1, 18):
                # 只看成功的方法
                if row[f"方法{i}_结果"]:
                    kill_val = row[f"方法{i}_杀号"]
                    if kill_val is None:
                        continue
                    # 解析出列表 nums
                    if isinstance(kill_val, str):
                        nums = [int(x) for x in kill_val.split(',') if x.strip().isdigit()]
                    elif isinstance(kill_val, (list, tuple)):
                        nums = kill_val
                    else:
                        nums = [kill_val]
                    # 对每个号码，记下是哪个方法
                    for num in nums:
                        if num in success_numbers_set:
                            number_sources.setdefault(num, []).append(f"方法{i}")

            # 把 dict 格式化
            # 比如 “5:(方法3,方法7); 12:(方法1); 17:(方法2,方法5,方法9)”
            source_str = '; '.join(
                f"{num}:({','.join(number_sources[num])})"
                for num in sorted(number_sources)
            )
            all_numbers = set(range(1, 34))
            remaining_numbers_success = sorted(all_numbers - success_numbers_set)
            remaining_numbers_all = sorted(all_numbers - all_killed_numbers_set)
            total_kill_count = len(all_killed_numbers_set)

            sheet.cell(row=row_idx + 2, column=36).value = ','.join(map(str, success_numbers))
            sheet.cell(row=row_idx + 2, column=37).value = f"{success_count}/{total_kill_count}"
            sheet.cell(row=row_idx + 2, column=38).value = ','.join(map(str, remaining_numbers_success))
            sheet.cell(row=row_idx + 2, column=39).value = ','.join(map(str, remaining_numbers_all))
            sheet.cell(row=row_idx + 2, column=40).value = source_str

        # 统计所有期次的杀号个数以及所有方法的成功次数
        kill_counts = []  # 每期成功杀号号码的个数
        overall_success_count = 0  # 所有方法成功次数累计
        all_methods_success_count = 0  # 每期所有方法均成功的次数
        for idx, row in results_df.iterrows():
            success_numbers_set = set()
            all_success = True
            for i in range(1, total_methods + 1):
                result_field = f"方法{i}_结果"
                kill_field = f"方法{i}_杀号"
                if row[result_field]:
                    overall_success_count += 1
                    kill_value = row[kill_field]
                    if kill_value is not None:
                        if isinstance(kill_value, str):
                            nums = [int(x.strip()) for x in kill_value.split(',') if x.strip().isdigit()]
                        elif isinstance(kill_value, (list, tuple)):
                            nums = kill_value
                        else:
                            nums = [kill_value]
                        success_numbers_set.update(nums)
                else:
                    all_success = False
            kill_counts.append(len(success_numbers_set))
            if all_success:
                all_methods_success_count += 1

        max_kill_count = max(kill_counts) if kill_counts else 0
        min_kill_count = min(kill_counts) if kill_counts else 0
        avg_kill_count = sum(kill_counts) / len(kill_counts) if kill_counts else 0
        combined_success_rate = overall_success_count / (total_methods * len(results_df)) if len(results_df) > 0 else 0

        # 新增统计：成功杀号个数达到某个阈值的次数
        kill_ge_10_count = sum(1 for count in kill_counts if count >= 10)
        kill_ge_15_count = sum(1 for count in kill_counts if count >= 15)
        kill_ge_20_count = sum(1 for count in kill_counts if count >= 20)
        # 写入成功率统计
        row_offset = len(results_df) + 3
        sheet.cell(row=row_offset, column=1).value = "杀号方法成功率统计"
        for i, (method, rate) in enumerate(success_rates.items(), 1):
            sheet.cell(row=row_offset + i, column=1).value = method
            cell = sheet.cell(row=row_offset + i, column=2)
            cell.value = f"{rate:.2%}"
            if rate >= 0.8:
                cell.fill = success_fill
            elif rate < 0.5:
                cell.fill = fail_fill

        # 写入下一期预测部分
        row_offset = row_offset + len(success_rates) + 2
        next_period = results_df["期数"].iloc[-1] + 1
        sheet.cell(row=row_offset, column=1).value = f"{next_period}期红球杀号预测"
        # 写入预测表头：方法、杀号
        sheet.cell(row=row_offset + 1, column=1).value = "方法"
        sheet.cell(row=row_offset + 1, column=2).value = "杀号"
        # 写入各方法预测结果，同时计算所有方法预测的杀号号码并集
        union_next_killed = set()
        for i, (method, numbers) in enumerate(next_predictions.items(), 1):
            row_index = row_offset + 1 + i
            sheet.cell(row=row_index, column=1).value = method
            # 如果预测数据不是列表，则转换为列表
            if not isinstance(numbers, (list, tuple)):
                numbers = [numbers]
            kill_numbers_str = ','.join(map(str, numbers))
            sheet.cell(row=row_index, column=2).value = kill_numbers_str
            union_next_killed.update(numbers)

        # ------------------------
        # 修改部分：根据各方法成功率，为号码1-33计算风险分数，并划分优先级
        #
        # 对于号码 1～33，遍历每个方法。如果该方法的预测中包含该号码，则加上该方法的成功率，
        # 得到该号码的风险分数（风险越低，说明被高成功率方法杀号的机会越低，则优先级越高）。
        risk_scores = {}
        for number in range(1, 34):
            risk = 0.0
            for method, numbers in next_predictions.items():
                # 如果不是列表，则转换为列表
                if not isinstance(numbers, (list, tuple)):
                    numbers = [numbers]
                if number in numbers:
                    risk += success_rates.get(method, 0)
            risk_scores[number] = risk

        # 根据风险分数划分优先级（阈值可根据实际情况调整）
        # 例如：风险分数为0 的归为【优先级1】；大于0但<=0.5 为【优先级2】；
        # 大于0.5但<=1.0 为【优先级3】，大于1.0 为【优先级4】
        priority_numbers = {"优先级1": [], "优先级2": [], "优先级3": [], "优先级4": []}
        for number, risk in risk_scores.items():
            if risk == 0:
                priority_numbers["优先级1"].append(number)
            elif risk <= 0.5:
                priority_numbers["优先级2"].append(number)
            elif risk <= 1.0:
                priority_numbers["优先级3"].append(number)
            else:
                priority_numbers["优先级4"].append(number)
        for key in priority_numbers:
            priority_numbers[key] = sorted(priority_numbers[key])

        # 将分优先级的剩余可用号码写入Excel，
        # 注意：这里写入的信息包括各优先级及对应的号码列表
        row_index = row_offset + 1 + len(next_predictions) + 1
        sheet.cell(row=row_index, column=1).value = "剩余可用号码（按优先级）"
        row_index += 1
        for priority, numbers in priority_numbers.items():
            sheet.cell(row=row_index, column=1).value = priority
            sheet.cell(row=row_index, column=2).value = ','.join(map(str, numbers))
            row_index += 1
        # ------------------------

        # 写入额外统计项目
        stat_start = row_index + 2
        sheet.cell(row=stat_start, column=1).value = "最多杀号个数"
        sheet.cell(row=stat_start, column=2).value = max_kill_count

        sheet.cell(row=stat_start + 1, column=1).value = "最少杀号个数"
        sheet.cell(row=stat_start + 1, column=2).value = min_kill_count

        sheet.cell(row=stat_start + 2, column=1).value = "平均杀号个数"
        sheet.cell(row=stat_start + 2, column=2).value = avg_kill_count

        sheet.cell(row=stat_start + 3, column=1).value = "所有方法综合成功率"
        cell = sheet.cell(row=stat_start + 3, column=2)
        cell.value = f"{combined_success_rate:.2%}"
        if combined_success_rate >= 0.8:
            cell.fill = success_fill
        elif combined_success_rate < 0.5:
            cell.fill = fail_fill

        sheet.cell(row=stat_start + 4, column=1).value = "所有方法都成功的次数"
        sheet.cell(row=stat_start + 4, column=2).value = all_methods_success_count

        sheet.cell(row=stat_start + 5, column=1).value = "成功杀号10个以上的次数"
        sheet.cell(row=stat_start + 5, column=2).value = kill_ge_10_count

        sheet.cell(row=stat_start + 6, column=1).value = "成功杀号15个以上的次数"
        sheet.cell(row=stat_start + 6, column=2).value = kill_ge_15_count

        sheet.cell(row=stat_start + 7, column=1).value = "成功杀号20个以上的次数"
        sheet.cell(row=stat_start + 7, column=2).value = kill_ge_20_count

        # 保存工作簿
        workbook.save(excel_file)
        logging.info(f"红球杀号结果已写入Excel文件 '{excel_file}' 的 '红球杀号' 工作表")
        return True
    except Exception as e:
        logging.error(f"写入Excel时发生错误: {e}")
        return False



def run_time_series_prediction(excel_file, apply_blue_kill=True):
    sheet_name = "开奖结果"
    time_str = time.strftime("%Y年%m月%d日_%H:%M:%S")
    try:
        df = pd.read_excel(excel_file, sheet_name=sheet_name)
    except FileNotFoundError:
        logging.error(f"错误: 文件 '{excel_file}' 未找到。")
        return
    except ValueError:
        logging.error(f"错误: 工作表 '{sheet_name}' 不存在于 '{excel_file}' 中。")
        return

    # 按期数排序并去重
    df.sort_values(by="期数", ascending=True, inplace=True)
    initial_count = len(df)
    df.drop_duplicates(subset="期数", inplace=True)
    final_count = len(df)
    if initial_count != final_count:
        logging.warning(f"移除了 {initial_count - final_count} 个重复的期数。")
    df.reset_index(drop=True, inplace=True)
    logging.info("读取并整理数据，按期数升序排序并移除重复。")

    logging.info(f"读取完毕，共 {len(df)} 条记录。")
    latest_period = df["期数"].max()
    next_period = latest_period + 1
    logging.info(f"最新一期的期数是：{latest_period}")

    # 计算红球杀号方法的成功率，用于日志显示
    _, red_success_rates, _ = red_kill_methods(df)
    high_success_methods = [method for method, rate in red_success_rates.items() if rate > 0.7]
    
    if high_success_methods:
        logging.info(f"将应用成功率高于70%的红球杀号方法: {', '.join(high_success_methods)}")
    else:
        logging.info("没有成功率高于70%的红球杀号方法可应用")

    methods = ["ma", "es", "rf", "svr", "bayes", "arima", "lstm", "hybrid"]
    results_map = {}
    next_period = latest_period + 1

    def run_method(m):
        logging.info(f"开始使用方法 {m.upper()} 进行Preview，使用最近 {len(df)} 期数据。")
        buf, res = recommend_by_method(df, method=m, lookback=len(df), apply_blue_kill=apply_blue_kill)  # 使用全部数据，并应用杀号
        results_map[m] = (buf, res)
        logging.info(f"方法 {m.upper()} 预测完成。")

    # 使用线程池并行运行预测方法
    with ThreadPoolExecutor(max_workers=1) as executor:
        futures = [executor.submit(run_method, m) for m in methods]
        for future in as_completed(futures):
            pass  # 所有结果已存入 results_map


    # 写入Preview详情
    with open(details_file_path, 'w', encoding='utf-8') as f:  # 使用'w'模式而不是'a'模式
        f.write(f"预测详情\n\n{next_period}期Preview详情-{time_str}\n")
        f.write(f"使用数据：最近 {len(df)} 期\n\n")
        for m in methods:
            if m in results_map:  # 检查方法是否在results_map中
                buf, res = results_map[m]
                if buf is not None:
                    for line in buf:
                        print(line)
                        f.write(f"{line}\n")
                if res is not None:
                    print(f"方法 {m.upper()} 推荐: {res}")
                    f.write(f"方法 {m.upper()} 推荐: {res}\n")
            else:
                print(f"方法 {m.upper()} 未完成预测")
                f.write(f"方法 {m.upper()} 未完成预测\n")
    
    # 同时写入到preview详情.txt文件
    with open(preview_details_path, 'w', encoding='utf-8') as f:
        f.write(f"{next_period}期Preview详情-{time_str}\n")
        f.write(f"使用数据：最近 {len(df)} 期\n\n")
        for m in methods:
            if m in results_map:  # 检查方法是否在results_map中
                buf, res = results_map[m]
                if buf is not None:
                    for line in buf:
                        f.write(f"{line}\n")
                if res is not None:
                    f.write(f"方法 {m.upper()} 推荐: {res}\n")
            else:
                f.write(f"方法 {m.upper()} 未完成预测\n")

    # 生成综合推荐
    recommendations = generate_recommendations(results_map, num_recommendations=5)
    
    # 运行蓝球杀号方法分析
    logging.info("开始分析蓝球杀号方法...")
    blue_results_df, blue_success_rates, blue_next_predictions = blue_kill_methods(df)
    
    # 将蓝球杀号结果写入Excel
    write_blue_kill_results_to_excel(excel_file, blue_results_df, blue_success_rates, blue_next_predictions)
    
    # 运行红球杀号方法分析
    logging.info("开始分析红球杀号方法...")
    red_results_df, red_success_rates, red_next_predictions = red_kill_methods(df)
    
    # 将红球杀号结果写入Excel
    write_red_kill_results_to_excel(excel_file, red_results_df, red_success_rates, red_next_predictions)
    
    # 将蓝球杀号结果写入预测详情文件
    with open(details_file_path, 'a', encoding='utf-8') as f:
        f.write(f"\n\n蓝球杀号方法分析-{time_str}\n")
        f.write(f"各方法历史成功率:\n")
        for method, rate in blue_success_rates.items():
            success_msg = f"{method}成功率: {rate:.2%}"
            print(success_msg)
            f.write(f"{success_msg}\n")
        
        f.write(f"\n{next_period}期蓝球杀号预测:\n")
        for method, numbers in blue_next_predictions.items():
            prediction_msg = f"{method}杀号: {', '.join(map(str, numbers))}"
            print(prediction_msg)
            f.write(f"{prediction_msg}\n")
    
    # 将红球杀号结果写入预测详情文件
    with open(details_file_path, 'a', encoding='utf-8') as f:
        f.write(f"\n\n红球杀号方法分析-{time_str}\n")
        f.write(f"各方法历史成功率:\n")
        for method, rate in red_success_rates.items():
            success_msg = f"{method}成功率: {rate:.2%}"
            print(success_msg)
            f.write(f"{success_msg}\n")
        
        f.write(f"\n{next_period}期红球杀号预测:\n")
        for method, numbers in red_next_predictions.items():
            prediction_msg = f"{method}杀号: {', '.join(map(str, numbers))}"
            print(prediction_msg)
            f.write(f"{prediction_msg}\n")
    
    # 同时写入到preview详情.txt文件
    with open(preview_details_path, 'a', encoding='utf-8') as f:
        f.write(f"\n\n蓝球杀号方法分析-{time_str}\n")
        f.write(f"各方法历史成功率:\n")
        for method, rate in blue_success_rates.items():
            f.write(f"{method}成功率: {rate:.2%}\n")
        
        f.write(f"\n{next_period}期蓝球杀号预测:\n")
        for method, numbers in blue_next_predictions.items():
            f.write(f"{method}杀号: {', '.join(map(str, numbers))}\n")
    
    # 将红球杀号结果写入preview详情.txt文件
    with open(preview_details_path, 'a', encoding='utf-8') as f:
        f.write(f"\n\n红球杀号方法分析-{time_str}\n")
        f.write(f"各方法历史成功率:\n")
        for method, rate in red_success_rates.items():
            f.write(f"{method}成功率: {rate:.2%}\n")
        
        f.write(f"\n{next_period}期红球杀号预测:\n")
        for method, numbers in red_next_predictions.items():
            f.write(f"{method}杀号: {', '.join(map(str, numbers))}\n")
    
    # 写入摘要
    with open(summary_file_path, 'w', encoding='utf-8') as file:  # 使用'w'模式而不是'a'模式
        file.write(f"预测摘要\n\n{next_period}期预测摘要-{time_str}\n")
        file.write(f"使用数据：最近 {len(df)} 期\n")
        
        # 写入应用的红球杀号方法信息
        if high_success_methods:
            file.write(f"\n应用成功率高于70%的红球杀号方法: {', '.join(high_success_methods)}\n")
            file.write("这些方法的杀号结果已应用于红球预测\n")
        else:
            file.write("\n没有成功率高于70%的红球杀号方法可应用\n")
        
        # 写入蓝球杀号方法摘要
        file.write("\n蓝球杀号方法分析:\n")
        for method, rate in blue_success_rates.items():
            file.write(f"{method}成功率: {rate:.2%}\n")
        
        file.write(f"\n{next_period}期蓝球杀号预测:\n")
        for method, numbers in blue_next_predictions.items():
            file.write(f"{method}杀号: {', '.join(map(str, numbers))}\n")
        
        # 收集所有被杀的蓝球号码
        all_killed_blue_numbers = set()
        for numbers in blue_next_predictions.values():
            all_killed_blue_numbers.update(numbers)
        available_blue_numbers = [n for n in range(1, 17) if n not in all_killed_blue_numbers]
        file.write(f"\n综合杀号后剩余可选蓝球: {', '.join(map(str, sorted(available_blue_numbers)))}\n")
        
        # 写入红球杀号方法摘要
        file.write("\n红球杀号方法分析:\n")
        for method, rate in red_success_rates.items():
            file.write(f"{method}成功率: {rate:.2%}\n")
        
        file.write(f"\n{next_period}期红球杀号预测:\n")
        for method, numbers in red_next_predictions.items():
            file.write(f"{method}杀号: {', '.join(map(str, numbers))}\n")
        
        # 收集所有被杀的红球号码
        all_killed_red_numbers = set()
        for numbers in red_next_predictions.values():
            all_killed_red_numbers.update(numbers)
        available_red_numbers = [n for n in range(1, 34) if n not in all_killed_red_numbers]
        file.write(f"\n综合杀号后剩余可选红球: {', '.join(map(str, sorted(available_red_numbers)))}\n")
        
        # 写入各方法推荐
        file.write("\n各方法推荐号码:\n")
        red_counts = Counter()
        blue_counts = Counter()
        
        for method, (_, res) in results_map.items():
            if res is None:
                continue
            reds, blue = res
            for r in reds:
                red_counts[r] += 1
            blue_counts[blue] += 1
            file.write(f"方法 {method.upper()} 推荐: 红球 {reds}, 蓝球 {blue}\n")
        
        # 写入5注推荐号码组合
        file.write("\n推荐5注号码组合:\n")
        # 创建两个计数器，一个用于原始推荐，一个用于调整后的推荐
        original_red_counts = Counter()
        adjusted_red_counts = Counter()
        original_blue_counts = Counter()
        adjusted_blue_counts = Counter()
        
        for i, (reds, blue) in enumerate(recommendations, 1):
            # 记录原始推荐的计数
            for r in reds:
                original_red_counts[r] += 1
            original_blue_counts[blue] += 1
            
            # 检查推荐的红球是否在被杀号列表中
            killed_reds = [r for r in reds if r in all_killed_red_numbers]
            # 检查蓝球是否在被杀号列表中
            blue_killed = blue in all_killed_blue_numbers
            
            if killed_reds and available_red_numbers:
                # 找到出现次数最多的未被杀的红球来替换被杀的红球
                available_reds_count = {r: red_counts.get(r, 0) for r in available_red_numbers}
                # 替换被杀的红球
                new_reds = [r for r in reds if r not in all_killed_red_numbers]
                while len(new_reds) < 6:
                    # 找出未被选择且未被杀的红球中出现次数最多的
                    available_for_selection = [r for r in available_red_numbers if r not in new_reds]
                    if not available_for_selection:
                        break
                    new_red = max(available_for_selection, key=lambda r: available_reds_count.get(r, 0))
                    new_reds.append(new_red)
                new_reds = sorted(new_reds)
                
                # 记录调整后推荐的计数
                for r in new_reds:
                    adjusted_red_counts[r] += 1
                
                file.write(f"推荐{i}: 红球 {reds} (部分被杀) -> 调整为红球 {new_reds}, 蓝球 {blue}")
            else:
                # 如果没有被杀的红球，直接使用原始推荐
                # 原始和调整后的计数相同
                for r in reds:
                    adjusted_red_counts[r] += 1
                file.write(f"推荐{i}: 红球 {reds}, 蓝球 {blue}")
            
            # 检查蓝球是否被杀
            if blue_killed and available_blue_numbers:
                available_blues_count = {b: blue_counts.get(b, 0) for b in available_blue_numbers}
                new_blue = max(available_blues_count.items(), key=lambda x: x[1])[0] if available_blues_count else available_blue_numbers[0]
                
                # 记录调整后蓝球的计数
                adjusted_blue_counts[new_blue] += 1
                
                file.write(f" (被杀) -> 调整为蓝球 {new_blue}\n")
            else:
                # 如果蓝球没有被杀，原始和调整后的计数相同
                adjusted_blue_counts[blue] += 1
                file.write("\n")
        
        # 统计原始推荐中被选中次数最多的号码
        file.write("\n[原始推荐统计] 被选中次数最多的红球号码\n")
        sorted_original_red = sorted(original_red_counts.items(), key=lambda x: x[1], reverse=True)
        for i in range(min(10, len(sorted_original_red))):
            num, cnt = sorted_original_red[i]
            file.write(f"红球 {num} 出现 {cnt} 次\n")
        
        file.write("\n[原始推荐统计] 被选中次数最多的蓝球号码\n")
        sorted_original_blue = sorted(original_blue_counts.items(), key=lambda x: x[1], reverse=True)
        for i in range(min(5, len(sorted_original_blue))):
            num, cnt = sorted_original_blue[i]
            file.write(f"蓝球 {num} 出现 {cnt} 次\n")
        
        # 找出原始推荐中没有被任何方法选中的红球号码
        unselected_original_reds = [num for num in range(1, 34) if num not in original_red_counts]
        if unselected_original_reds:
            file.write("\n[原始推荐统计] 未被任何方法选中的红球号码\n")
            for r in unselected_original_reds:
                file.write(f"红球 {r} 未被选中\n")
        
        # 统计调整后推荐中被选中次数最多的号码
        file.write("\n[调整后推荐统计] 被选中次数最多的红球号码\n")
        sorted_adjusted_red = sorted(adjusted_red_counts.items(), key=lambda x: x[1], reverse=True)
        for i in range(min(10, len(sorted_adjusted_red))):
            num, cnt = sorted_adjusted_red[i]
            file.write(f"红球 {num} 出现 {cnt} 次\n")
        
        file.write("\n[调整后推荐统计] 被选中次数最多的蓝球号码\n")
        sorted_adjusted_blue = sorted(adjusted_blue_counts.items(), key=lambda x: x[1], reverse=True)
        for i in range(min(5, len(sorted_adjusted_blue))):
            num, cnt = sorted_adjusted_blue[i]
            file.write(f"蓝球 {num} 出现 {cnt} 次\n")
        
        # 找出调整后推荐中没有被任何方法选中的红球号码
        unselected_adjusted_reds = [num for num in range(1, 34) if num not in adjusted_red_counts]
        if unselected_adjusted_reds:
            file.write("\n[调整后推荐统计] 未被任何方法选中的红球号码\n")
            for r in unselected_adjusted_reds:
                file.write(f"红球 {r} 未被选中\n")

        # 打印终端摘要
        print("\n>>> [数据统计] <<<")
        print(f"使用数据：最近 {len(df)} 期")
        print(f"最新一期：{latest_period}")
        print(f"预测期数：{next_period}")
        
        print("\n>>> [蓝球杀号] <<<")
        for method, rate in blue_success_rates.items():
            print(f"{method}成功率: {rate:.2%}")
        
        print(f"\n{next_period}期蓝球杀号预测:")
        for method, numbers in blue_next_predictions.items():
            print(f"{method}杀号: {', '.join(map(str, numbers))}")
        
        print(f"\n综合杀号后剩余可选蓝球: {', '.join(map(str, sorted(available_blue_numbers)))}")
        
        print("\n>>> [红球杀号] <<<")
        for method, rate in red_success_rates.items():
            print(f"{method}成功率: {rate:.2%}")
        
        print(f"\n{next_period}期红球杀号预测:")
        for method, numbers in red_next_predictions.items():
            print(f"{method}杀号: {', '.join(map(str, numbers))}")
        
        print(f"\n综合杀号后剩余可选红球: {', '.join(map(str, sorted(available_red_numbers)))}")
        
        print("\n>>> [推荐号码] 5注组合 <<<")
        for i, (reds, blue) in enumerate(recommendations, 1):
            # 检查推荐的红球是否在被杀号列表中
            killed_reds = [r for r in reds if r in all_killed_red_numbers]
            # 检查蓝球是否在被杀号列表中
            blue_killed = blue in all_killed_blue_numbers
            
            if killed_reds and available_red_numbers:
                # 找到出现次数最多的未被杀的红球来替换被杀的红球
                available_reds_count = {r: red_counts.get(r, 0) for r in available_red_numbers}
                new_reds = reds.copy()
                for killed_red in killed_reds:
                    # 找到一个未被使用且未被杀的红球
                    potential_replacements = [r for r in available_red_numbers if r not in new_reds]
                    if potential_replacements:
                        # 按出现频率排序
                        replacement = max(potential_replacements, key=lambda r: available_reds_count.get(r, 0))
                        # 替换被杀的红球
                        new_reds[new_reds.index(killed_red)] = replacement
                
                if blue_killed and available_blue_numbers:
                    # 找到出现次数最多的未被杀的蓝球
                    available_blues_count = {b: blue_counts.get(b, 0) for b in available_blue_numbers}
                    new_blue = max(available_blues_count.items(), key=lambda x: x[1])[0] if available_blues_count else available_blue_numbers[0]
                    print(f"推荐{i}: 红球 {reds} (部分被杀) -> 调整为红球 {new_reds}, 蓝球 {blue} (被杀) -> 调整为蓝球 {new_blue}")
                else:
                    print(f"推荐{i}: 红球 {reds} (部分被杀) -> 调整为红球 {new_reds}, 蓝球 {blue}")
            elif blue_killed and available_blue_numbers:
                # 找到出现次数最多的未被杀的蓝球
                available_blues_count = {b: blue_counts.get(b, 0) for b in available_blue_numbers}
                new_blue = max(available_blues_count.items(), key=lambda x: x[1])[0] if available_blues_count else available_blue_numbers[0]
                print(f"推荐{i}: 红球 {reds}, 蓝球 {blue} (被杀) -> 调整为蓝球 {new_blue}")
            else:
                print(f"推荐{i}: 红球 {reds}, 蓝球 {blue}")
        
        print("\n>>> [推荐号码] 生成完毕 <<<\n")

def main(excel_file):
    # 默认应用蓝球杀号方法，除非指定了--no-kill参数
    apply_blue_kill = not args.no_kill
    if apply_blue_kill:
        logging.info("将应用蓝球杀号方法来优化预测结果")
    else:
        logging.info("不应用蓝球杀号方法")
    
    run_time_series_prediction(excel_file, apply_blue_kill=apply_blue_kill)

if __name__ == "__main__":
    main(args.excel_file)
